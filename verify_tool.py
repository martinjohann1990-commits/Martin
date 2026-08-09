#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verifikation des erzeugten Cashflow-Tools.

Teil A  Nachrechnung des Prognosemodells in reinem Python mit denselben
        Beispieldaten, um die Wirkung der Beispieldaten zu pruefen.
Teil B  Statische Analyse der erzeugten .xlsx: Blattbezuege, benannte
        Bereiche, Ueberlappungen verbundener Zellen, Formelabdeckung.
"""

import re
import sys
from datetime import date, timedelta

import openpyxl
from openpyxl.utils import range_boundaries

import build_cashflow_tool as B


# ---------------------------------------------------------------------------
# Teil A: Modell-Nachrechnung
# ---------------------------------------------------------------------------

def month_day(d, day):
    """DATE(YEAR(d), MONTH(d), day) - Tage 1..28 sind immer gueltig."""
    return date(d.year, d.month, min(day, 28))


def simulate(delay_all, delay_cust, loss, customer, today=None):
    today = today or date.today()
    start = today - timedelta(days=today.weekday())
    weeks = [(start + timedelta(days=7 * i), start + timedelta(days=7 * i + 6))
             for i in range(12)]

    rows = []
    balance_base = balance_scn = B.OPEN_BALANCE
    for ws_, we in weeks:
        inflow_base = inflow_scn = 0.0
        for cust, _no, amount, d_inv, terms, paid in B.INVOICES:
            if paid:
                continue
            due = today + timedelta(days=d_inv) + timedelta(days=terms)
            base_d = max(due, start)
            scn_d = base_d + timedelta(days=delay_all +
                                       (delay_cust if cust == customer else 0))
            if ws_ <= base_d <= we:
                inflow_base += amount
            if ws_ <= scn_d <= we:
                inflow_scn += amount * (1 - (loss if cust == customer else 0))

        out_fixed = 0.0
        for _n, _c, amount, day in B.FIXED_COSTS:
            hits = {month_day(ws_, day), month_day(we, day)}
            if any(ws_ <= h <= we for h in hits):
                out_fixed += amount

        out_var = sum(a for _n, _c, a, off in B.VAR_COSTS
                      if ws_ <= today + timedelta(days=off) <= we)

        out_total = out_fixed + out_var
        balance_base += inflow_base - out_total
        balance_scn += inflow_scn - out_total
        code = 2 if balance_scn < B.MIN_RESERVE else (
            1 if balance_scn < B.MIN_RESERVE * B.YELLOW_FACT else 0)
        rows.append((ws_, inflow_base, inflow_scn, out_total,
                     balance_base, balance_scn, code))
    return rows


def report_model():
    print("=" * 88)
    print("TEIL A  Nachrechnung des Prognosemodells   (heute = %s)" % date.today())
    print("=" * 88)
    for label, args in (("BASIS  (keine Simulation)", (0, 0, 0.0, None)),
                        ("SZENARIO (Voreinstellung der Demo)",
                         (B.SCN_DELAY_ALL, B.SCN_DELAY_CUST, B.SCN_LOSS, B.SCN_CUSTOMER))):
        rows = simulate(*args)
        print("\n%s" % label)
        print("  KW-Start     Einzahlung   Auszahlung   Saldo Basis  Saldo Szen.  Ampel")
        marks = {0: "gruen", 1: "GELB", 2: "ROT"}
        for ws_, inb, ins, out, bb, bs, code in rows:
            print("  %s  %11.0f  %11.0f  %11.0f  %11.0f  %s"
                  % (ws_.strftime("%d.%m.%Y"), ins if args[0] or args[1] else inb,
                     -out, bb, bs, marks[code]))
        codes = [r[6] for r in rows]
        print("  -> Tiefststand Basis %.0f | Szenario %.0f | Ampelverteilung "
              "gruen=%d gelb=%d rot=%d"
              % (min(r[4] for r in rows), min(r[5] for r in rows),
                 codes.count(0), codes.count(1), codes.count(2)))
    return simulate(B.SCN_DELAY_ALL, B.SCN_DELAY_CUST, B.SCN_LOSS, B.SCN_CUSTOMER)


# ---------------------------------------------------------------------------
# Teil B: Statische Analyse der Datei
# ---------------------------------------------------------------------------

EXCEL_FUNCS = {
    "IF", "AND", "OR", "NOT", "SUM", "SUMIFS", "COUNTIF", "COUNTIFS", "INDEX",
    "MATCH", "MIN", "MAX", "N", "DATE", "YEAR", "MONTH", "DAY", "TODAY",
    "WEEKDAY", "ISOWEEKNUM", "IFERROR", "ROW", "COLUMN", "EOMONTH", "ABS",
    "TEXT", "VALUE", "ROUND", "LEFT", "RIGHT", "MID", "TRIM", "EXACT",
}


def check_file(path, model_rows):
    print("\n" + "=" * 88)
    print("TEIL B  Statische Analyse von %s" % path)
    print("=" * 88)

    wb = openpyxl.load_workbook(path)
    sheets = set(wb.sheetnames)
    names = set(wb.defined_names.keys())
    problems = []

    sheet_ref = re.compile(r"'([^']+)'!|(?<![A-Za-z0-9_'!])([A-Za-z_][A-Za-z0-9_]*)!")
    # Namenskandidaten: keine Zellbezuege (A1, AB$25) und keine Funktionsaufrufe
    token = re.compile(r"(?<![A-Za-z0-9_$!'.])([A-Z][A-Z0-9_]{1,})(?![A-Za-z0-9_(!$])")
    cellref = re.compile(r"^\$?[A-Z]{1,3}\$?[0-9]{1,7}$")

    formula_count = 0
    referenced_sheets, referenced_names = set(), set()

    for ws in wb.worksheets:
        # 1) Ueberlappende verbundene Zellbereiche
        merged = [range_boundaries(str(m)) for m in ws.merged_cells.ranges]
        for i, a in enumerate(merged):
            for b in merged[i + 1:]:
                if not (a[2] < b[0] or b[2] < a[0] or a[3] < b[1] or b[3] < a[1]):
                    problems.append("%s: verbundene Bereiche ueberlappen: %s / %s"
                                    % (ws.title, a, b))

        # 2) Formeln einsammeln und Bezuege pruefen
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                formula_count += 1
                for q, plain in sheet_ref.findall(v):
                    referenced_sheets.add(q or plain)
                for tok in token.findall(v):
                    if tok in EXCEL_FUNCS or cellref.match(tok):
                        continue
                    referenced_names.add(tok)

    for s in sorted(referenced_sheets):
        if s not in sheets:
            problems.append("Formelbezug auf unbekanntes Blatt: %r" % s)
    for n in sorted(referenced_names):
        if n not in names:
            problems.append("Formel nutzt undefinierten Namen: %r" % n)

    # 3) Ziele der benannten Bereiche muessen existieren und belegt sein
    for name in sorted(names):
        dn = wb.defined_names[name]
        ref = dn.attr_text
        sheet = ref.split("!")[0].strip("'")
        if sheet not in sheets:
            problems.append("Name %s zeigt auf unbekanntes Blatt %r" % (name, sheet))
            continue
        coord = ref.split("!")[1].replace("$", "")
        cell = wb[sheet][coord]
        if cell.value is None:
            problems.append("Name %s -> %s!%s ist leer" % (name, sheet, coord))

    # 4) Uebersetzungstabelle vollstaendig und eindeutig
    cfg = wb[B.SH_CFG]
    keys = [cfg.cell(row=r, column=1).value
            for r in range(B.TR_FIRST, B.TR_LAST + 1)]
    if len(keys) != len(set(keys)):
        problems.append("Doppelte Uebersetzungsschluessel vorhanden")
    for r in range(B.TR_FIRST, B.TR_LAST + 1):
        for c, lang in ((2, "DE"), (3, "EN")):
            if not cfg.cell(row=r, column=c).value:
                problems.append("Uebersetzung fehlt: Zeile %d, %s" % (r, lang))

    # 5) Jeder in Formeln verwendete Schluessel existiert in der Tabelle
    used = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "MATCH(\"" in cell.value:
                    used.update(re.findall(r'MATCH\("([a-z0-9._]+)"', cell.value))
    missing = sorted(used - set(keys))
    if missing:
        problems.append("Verwendete, aber nicht definierte Schluessel: %s" % missing)

    # 6) Rechenkern belegt
    for col in "MNOPQRSTUVWXY":
        for r in range(5, 17):
            if cfg["%s%d" % (col, r)].value is None:
                problems.append("Rechenkern leer: Config!%s%d" % (col, r))
                break

    print("Blaetter             : %s" % ", ".join(wb.sheetnames))
    print("Formeln gesamt       : %d" % formula_count)
    print("Benannte Bereiche    : %d (%s)" % (len(names), ", ".join(sorted(names))))
    print("Uebersetzungspaare   : %d" % len(keys))
    print("Verwendete Textkeys  : %d" % len(used))
    print("Config-Blatt versteckt: %s" % (wb[B.SH_CFG].sheet_state == "hidden"))
    print("Diagramme            : %s"
          % {ws.title: len(ws._charts) for ws in wb.worksheets if ws._charts})
    print("Bedingte Formate     : %s"
          % {ws.title: len(list(ws.conditional_formatting)) for ws in wb.worksheets
             if list(ws.conditional_formatting)})

    print("\nBEFUND: %s" % ("keine Probleme gefunden" if not problems
                            else "%d Problem(e)" % len(problems)))
    for p in problems:
        print("  ! %s" % p)
    return problems


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "Cashflow-Fruehwarnsystem.xlsx"
    rows = report_model()
    issues = check_file(path, rows)
    sys.exit(1 if issues else 0)
