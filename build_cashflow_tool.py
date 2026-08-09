#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cashflow-Fruehwarnsystem / Cashflow Early Warning System
=========================================================
Erzeugt eine zweisprachige (DE/EN), formelbasierte .xlsx-Datei ohne Makros,
kompatibel mit Microsoft Excel und Google Sheets.

Aufruf:  python3 build_cashflow_tool.py [ausgabedatei.xlsx]
"""

import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.numbers import NumberFormat
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont,
)

# ---------------------------------------------------------------------------
# 1. DESIGN-SYSTEM
# ---------------------------------------------------------------------------

PETROL      = "103C4B"   # Kopfband, Titel, Tabellenkoepfe
PETROL_SOFT = "1B4F63"   # Sekundaerband
TEAL        = "1B7F8C"   # Primaerakzent, Chartlinie Basis
SAGE        = "2FA88F"   # Sekundaerakzent, Chartlinie Szenario
CANVAS      = "F4F7F9"   # Blattflaeche
CARD        = "FFFFFF"   # Kachelflaeche
HAIRLINE    = "DDE6EC"   # dezente Rahmen
GRIDLINE    = "E9EFF3"   # Chart-Gitternetz
ZEBRA       = "FAFCFD"   # Zeilenstreifen
TEXT        = "1E2A32"   # Primaertext
MUTED       = "6B7C8A"   # Sekundaertext
INPUT_BG    = "FFFDF5"   # Eingabefelder (warmer Ton)
INPUT_BD    = "E4D9B8"   # Rahmen Eingabefelder
CALC_BG     = "F2F6F8"   # berechnete Felder

GREEN_BG, GREEN_FG = "E6F4EE", "1E8A5F"
AMBER_BG, AMBER_FG = "FDF4DE", "A9761A"
RED_BG,   RED_FG   = "FAE8E6", "B8402F"

FONT = "Segoe UI"        # faellt in Google Sheets sauber auf Arial zurueck

# Zahlenformate: Basis = deutsch (Symbol hinten), EN-Variante via bedingter
# Formatierung (Symbol vorn).  Dezimaltrennzeichen folgen der Locale des
# Anwenders - das ist eine Eigenschaft von Excel, nicht der Datei.
FMT_EUR_DE   = '#,##0.00\\ "€"'
FMT_EUR_EN   = '"€"\\ #,##0.00'
FMT_EUR0_DE  = '#,##0\\ "€"'
FMT_EUR0_EN  = '"€"\\ #,##0'
FMT_EUR0Z_DE = '#,##0\\ "€";-#,##0\\ "€";"–"'
FMT_EUR0Z_EN = '"€"\\ #,##0;-"€"\\ #,##0;"–"'
FMT_DATE_DE  = 'DD.MM.YYYY'
FMT_DATE_EN  = 'MM/DD/YYYY'
FMT_PCT      = '0%'

THIN = Side(style="thin", color=HAIRLINE)

# Blattnamen (Pipe ist in Excel und Google Sheets als Blattname zulaessig)
SH_DASH  = "Dashboard"
SH_INV   = "Rechnungen | Invoices"
SH_COST  = "Kosten | Costs"
SH_SCN   = "Szenario | Scenario"
SH_CFG   = "Config"

QINV  = "'" + SH_INV + "'"
QCOST = "'" + SH_COST + "'"
QSCN  = "'" + SH_SCN + "'"

# ---------------------------------------------------------------------------
# 2. UEBERSETZUNGSTABELLE
# ---------------------------------------------------------------------------

TRANSLATIONS = [
    # --- Allgemein -------------------------------------------------------
    ("app.title",        "Cashflow-Frühwarnsystem", "Cashflow Early Warning System"),
    ("app.sub",          "Liquiditätsplanung für Kleinunternehmer und Selbstständige",
                         "Liquidity planning for small businesses and freelancers"),
    ("ui.language",      "Sprache / Language", "Sprache / Language"),

    # --- Dashboard: Kennzahlen -------------------------------------------
    ("kpi.balance",      "Kontostand heute", "Cash balance today"),
    ("kpi.balance.sub",  "Startwert der Prognose", "starting value of the forecast"),
    ("kpi.recv",         "Offene Forderungen", "Open receivables"),
    ("kpi.recv.sub",     "unbezahlte Rechnungen", "unpaid invoices"),
    ("kpi.low",          "Tiefststand 12 Wochen", "Lowest point, 12 weeks"),
    ("kpi.low.sub",      "inklusive Szenario", "including scenario"),
    ("kpi.status",       "Frühwarn-Status", "Early warning status"),
    ("kpi.status.sub",   "gemessen an der Mindestreserve", "measured against the minimum reserve"),

    # --- Ampel ------------------------------------------------------------
    ("status.green",     "Alles im grünen Bereich", "All clear"),
    ("status.amber",     "Es wird eng", "Getting tight"),
    ("status.red",       "Liquiditätsengpass", "Cash shortfall"),
    ("alert.green",      "In den nächsten 12 Wochen bleibt Ihr Kontostand durchgehend über der Mindestreserve.",
                         "Your cash balance stays above the minimum reserve for the entire 12-week horizon."),
    ("alert.amber",      "Achtung: Ihr Kontostand nähert sich der Mindestreserve ab",
                         "Caution: your cash balance approaches the minimum reserve from"),
    ("alert.red",        "Warnung: Ihr Kontostand fällt unter die Mindestreserve ab",
                         "Warning: your cash balance drops below the minimum reserve from"),

    # --- Dashboard: Abschnitte -------------------------------------------
    ("sec.chart",        "Liquiditätsverlauf – nächste 12 Wochen", "Cash flow – next 12 weeks"),
    ("sec.table",        "Wochenübersicht", "Weekly overview"),
    ("sec.settings",     "Einstellungen", "Settings"),
    ("sec.help",         "So funktioniert es", "How it works"),

    # --- Wochentabelle ----------------------------------------------------
    ("th.week",          "Woche", "Week"),
    ("th.from",          "Woche ab", "Week from"),
    ("th.in",            "Einzahlungen", "Money in"),
    ("th.out",           "Auszahlungen", "Money out"),
    ("th.base",          "Saldo Basis", "Balance base"),
    ("th.scn",           "Saldo Szenario", "Balance scenario"),
    ("lbl.week",         "KW", "Wk"),

    # --- Einstellungen ----------------------------------------------------
    ("set.openbal",      "Startkontostand", "Opening cash balance"),
    ("set.mincash",      "Mindestreserve", "Minimum reserve"),
    ("set.yellow",       "Gelb-Schwelle (× Reserve)", "Amber threshold (× reserve)"),
    ("set.start",        "Prognosestart (Montag)", "Forecast start (Monday)"),
    ("set.vat",          "Umsatzsteuer aktiv?", "VAT active?"),
    ("set.vatday",       "USt-Zahltag im Monat", "VAT payment day of month"),
    ("set.hint",         "Kleinunternehmer nach §19 UStG lassen die Umsatzsteuer auf »Nein«.",
                         "Businesses exempt from VAT keep this setting on »No«."),

    # --- Diagramm ---------------------------------------------------------
    ("ser.base",         "Basis-Prognose", "Base forecast"),
    ("ser.scn",          "Szenario", "Scenario"),
    ("ser.min",          "Mindestreserve", "Minimum reserve"),

    # --- Hilfe ------------------------------------------------------------
    ("help.1",           "1.  Tragen Sie im Tab »Rechnungen« Ihre offenen Rechnungen ein: Kunde, Betrag, Rechnungsdatum und Zahlungsziel in Tagen.",
                         "1.  On the »Invoices« tab, enter your open invoices: customer, amount, invoice date and payment terms in days."),
    ("help.2",           "2.  Erfassen Sie im Tab »Kosten« Ihre monatlichen Fixkosten sowie geplante Einmalausgaben mit Datum.",
                         "2.  On the »Costs« tab, record your monthly fixed costs plus any planned one-off expenses with a date."),
    ("help.3",           "3.  Stellen Sie rechts Ihren aktuellen Kontostand und Ihre Mindestreserve ein – den Betrag, den Sie nie unterschreiten möchten.",
                         "3.  Set your current cash balance and minimum reserve on the right – the amount you never want to fall below."),
    ("help.4",           "4.  Die farbige Wochenmarkierung zeigt sofort, wann es eng wird: Grün unbedenklich, Gelb kritische Annäherung, Rot unter der Reserve.",
                         "4.  The coloured week marker shows immediately when things get tight: green is fine, amber is close, red is below the reserve."),
    ("help.5",           "5.  Im Tab »Szenario« simulieren Sie, was passiert, wenn Kunden später oder gar nicht zahlen – die Prognose reagiert sofort.",
                         "5.  On the »Scenario« tab, simulate what happens if customers pay late or not at all – the forecast reacts instantly."),
    ("foot.note",        "Alle Beträge in Euro. Diese Datei enthält keine Makros und funktioniert in Microsoft Excel wie in Google Sheets.",
                         "All amounts in euro. This file contains no macros and works in Microsoft Excel as well as Google Sheets."),

    # --- Rechnungen -------------------------------------------------------
    ("inv.title",        "Offene Rechnungen erfassen", "Record open invoices"),
    ("inv.hint",         "Weiße Felder ausfüllen, graue Felder rechnen sich selbst. Überfällige Rechnungen werden rot markiert und in der ersten Prognosewoche erwartet.",
                         "Fill in the white fields; the grey fields calculate themselves. Overdue invoices are marked red and are expected in the first forecast week."),
    ("inv.h.no",         "Nr.", "No."),
    ("inv.h.customer",   "Kunde", "Customer"),
    ("inv.h.invno",      "Rechnungs-Nr.", "Invoice no."),
    ("inv.h.amount",     "Betrag netto", "Amount net"),
    ("inv.h.vat",        "USt-Satz", "VAT rate"),
    ("inv.h.gross",      "Betrag brutto", "Gross amount"),
    ("inv.h.date",       "Rechnungsdatum", "Invoice date"),
    ("inv.h.terms",      "Zahlungsziel (Tage)", "Terms (days)"),
    ("inv.h.due",        "Fällig am", "Due date"),
    ("inv.h.status",     "Status", "Status"),
    ("inv.h.expdate",    "Erwarteter Eingang", "Expected receipt"),
    ("inv.h.expamt",     "Erwarteter Betrag", "Expected amount"),
    ("inv.total",        "Summe offener Forderungen", "Total open receivables"),
    ("inv.calc",         "berechnet – bitte nicht überschreiben", "calculated – please do not overwrite"),
    ("inv.scnnote",      "Diese beiden Spalten berücksichtigen Ihre Einstellungen im Tab »Szenario«.",
                         "These two columns reflect your settings on the »Scenario« tab."),

    # --- Kosten -----------------------------------------------------------
    ("cost.title",       "Kosten erfassen", "Record costs"),
    ("cost.fixed",       "Fixe monatliche Kosten", "Fixed monthly costs"),
    ("cost.fixed.hint",  "Kosten, die jeden Monat am gleichen Tag anfallen – Miete, Versicherungen, Abos, Leasing.",
                         "Costs that recur every month on the same day – rent, insurance, subscriptions, leasing."),
    ("cost.var",         "Variable und einmalige Kosten", "Variable and one-off costs"),
    ("cost.var.hint",    "Einzelausgaben mit konkretem Datum – Anschaffungen, Steuervorauszahlungen, Reisen, Kampagnen.",
                         "Individual expenses with a specific date – purchases, tax prepayments, travel, campaigns."),
    ("cost.h.name",      "Bezeichnung", "Description"),
    ("cost.h.cat",       "Kategorie", "Category"),
    ("cost.h.month",     "Betrag pro Monat", "Amount per month"),
    ("cost.h.day",       "Fällig am Tag (1–28)", "Due on day (1–28)"),
    ("cost.h.amount",    "Betrag", "Amount"),
    ("cost.h.date",      "Datum", "Date"),
    ("cost.sum.fixed",   "Summe Fixkosten pro Monat", "Total fixed costs per month"),
    ("cost.sum.var",     "Summe variable Kosten", "Total variable costs"),

    # --- Szenario ---------------------------------------------------------
    ("scn.title",        "Szenario-Simulation", "Scenario simulation"),
    ("scn.hint",         "Verändern Sie die drei Werte und beobachten Sie, wie Prognose und Ampel sofort reagieren – auch auf dem Dashboard.",
                         "Change the three values below and watch the forecast and warning lights react instantly – on the dashboard too."),
    ("scn.controls",     "Simulations-Regler", "Simulation controls"),
    ("scn.delay_all",    "Zahlungsverzug aller Kunden", "Payment delay, all customers"),
    ("scn.delay_all.h",  "Alle offenen Rechnungen gehen um so viele Tage später ein.",
                         "Every open invoice is received this many days later."),
    ("scn.customer",     "Kunde auswählen", "Select customer"),
    ("scn.customer.h",   "Für diesen einen Kunden gelten die beiden Werte darunter zusätzlich.",
                         "The two values below apply additionally to this one customer."),
    ("scn.delay_cust",   "Zusätzlicher Verzug dieses Kunden", "Additional delay for this customer"),
    ("scn.delay_cust.h", "Kommt zum allgemeinen Verzug oben hinzu.", "Added on top of the general delay above."),
    ("scn.loss",         "Forderungsausfall dieses Kunden", "Bad debt of this customer"),
    ("scn.loss.h",       "100 % bedeutet: Dieser Kunde zahlt gar nicht mehr.",
                         "100 % means: this customer does not pay at all."),
    ("scn.days",         "Tage", "days"),
    ("scn.reset",        "Alle drei Werte auf 0 setzen = keine Simulation, Szenario entspricht der Basis.",
                         "Set all three values to 0 = no simulation, the scenario equals the base case."),
    ("scn.impact",       "Auswirkung auf die Prognose", "Impact on the forecast"),
    ("scn.h.diff",       "Differenz", "Difference"),
    ("scn.low.base",     "Tiefststand Basis", "Lowest point, base"),
    ("scn.low.scn",      "Tiefststand Szenario", "Lowest point, scenario"),
    ("scn.gap",          "Verschlechterung", "Deterioration"),
    ("scn.firstwarn",    "Erste Warnwoche", "First warning week"),
    ("scn.none",         "keine", "none"),
    ("scn.chart",        "Basis gegen Szenario", "Base versus scenario"),

    # --- Werte / Dropdowns ------------------------------------------------
    ("val.open",         "offen", "open"),
    ("val.paid",         "bezahlt", "paid"),
    ("val.yes",          "Ja", "Yes"),
    ("val.no",           "Nein", "No"),
]

TR_FIRST = 6
TR_LAST  = TR_FIRST + len(TRANSLATIONS) - 1


def T(key):
    """Sprachabhaengige Textformel aus der Uebersetzungstabelle."""
    return ('=INDEX({c}!$B${a}:$C${b},MATCH("{k}",{c}!$A${a}:$A${b},0),LANG)'
            .format(c=SH_CFG, a=TR_FIRST, b=TR_LAST, k=key))


def Traw(key):
    """Wie T(), aber ohne fuehrendes '=' - zur Verkettung in groesseren Formeln."""
    return T(key)[1:]


# ---------------------------------------------------------------------------
# 3. BEISPIELDATEN
# ---------------------------------------------------------------------------

OPEN_BALANCE = 6800.00
MIN_RESERVE  = 5000.00
YELLOW_FACT  = 1.5
VAT_DAY      = 10

# (Kunde, Rechnungs-Nr., Netto, Rechnungsdatum relativ zu heute, Zahlungsziel, bezahlt?)
INVOICES = [
    ("Meier & Partner GmbH",    "RE-2601", 1850.00, -38, 30, False),
    ("Stadtwerke Lindau",       "RE-2602",  780.00, -31, 14, False),
    ("Kreativagentur Nord",     "RE-2603", 1240.00, -22, 30, False),
    ("Bauer Consulting KG",     "RE-2604", 4500.00, -18, 45, False),
    ("Hotel Seeblick",          "RE-2605",  640.00, -14, 14, False),
    ("TechnoPlus AG",           "RE-2606", 8400.00, -11, 60, False),
    ("Praxis Dr. Winter",       "RE-2607",  980.00,  -9, 30, False),
    ("Grünfeld Landschaftsbau", "RE-2608", 1150.00,  -6, 21, False),
    ("Meier & Partner GmbH",    "RE-2609", 2600.00,  -3, 30, False),
    ("Café Sonnenblume",        "RE-2610",  520.00,   1, 14, False),
    ("TechnoPlus AG",           "RE-2611", 5200.00,   4, 45, False),
    ("Schmidt Immobilien",      "RE-2612", 3100.00,   8, 30, False),
    ("Kreativagentur Nord",     "RE-2598", 1450.00, -52, 30, True),
    ("Hotel Seeblick",          "RE-2596",  780.00, -58, 14, True),
    ("Bauer Consulting KG",     "RE-2594", 3200.00, -66, 30, True),
]

# (Bezeichnung, Kategorie, Betrag pro Monat, Faelligkeitstag)
FIXED_COSTS = [
    ("Miete Büro / Office rent",              "Raumkosten / Premises",   850.00,  1),
    ("Krankenversicherung / Health insurance","Vorsorge / Insurance",    780.00,  1),
    ("Altersvorsorge / Pension plan",         "Vorsorge / Insurance",    300.00,  1),
    ("Leasing Fahrzeug / Car leasing",        "Fahrzeug / Vehicle",      320.00,  5),
    ("Software-Abos / Software subscriptions","IT / IT",                 165.00, 10),
    ("Betriebshaftpflicht / Liability cover", "Vorsorge / Insurance",     95.00, 12),
    ("Telefon & Internet / Phone & internet", "IT / IT",                  75.00, 15),
    ("Steuerberatung / Tax adviser",          "Beratung / Advisory",     180.00, 20),
    ("Buchhaltungstool / Accounting tool",    "IT / IT",                  39.00, 25),
]

# (Bezeichnung, Kategorie, Betrag, Datum relativ zu heute)
VAR_COSTS = [
    ("Reisekosten Kundentermin / Client travel",   "Reisen / Travel",       430.00,  9),
    ("Steuervorauszahlung / Tax prepayment",       "Steuern / Taxes",      3200.00, 12),
    ("Weiterbildung Seminar / Training seminar",   "Personal / People",     890.00, 17),
    ("Neue Hardware / New hardware",               "IT / IT",              1950.00, 26),
    ("Marketing-Kampagne / Marketing campaign",    "Marketing / Marketing",1250.00, 38),
    ("Messeauftritt / Trade fair booth",           "Marketing / Marketing",2650.00, 58),
]

SCN_DELAY_ALL  = 14
SCN_CUSTOMER   = "TechnoPlus AG"
SCN_DELAY_CUST = 21
SCN_LOSS       = 0.00

# ---------------------------------------------------------------------------
# 4. HILFSFUNKTIONEN FUER DAS LAYOUT
# ---------------------------------------------------------------------------


def box(ws, c1, r1, c2, r2, fill=None, border=HAIRLINE):
    """Zeichnet eine Kachel: Flaechenfuellung plus dezenter Aussenrahmen."""
    side = Side(style="thin", color=border) if border else None
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if side:
                cell.border = Border(
                    left=side if c == c1 else None,
                    right=side if c == c2 else None,
                    top=side if r == r1 else None,
                    bottom=side if r == r2 else None,
                )


def put(ws, coord, value, font=None, fill=None, align=None, fmt=None, border=None):
    cell = ws[coord]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border
    return cell


def paint(ws, c1, r1, c2, r2, fill):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=fill)


def canvas(ws, max_col, max_row):
    """Legt die Blattflaeche in Canvas-Farbe an und blendet das Gitternetz aus."""
    ws.sheet_view.showGridLines = False
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CANVAS)


def header_band(ws, c1, c2, title_key, sub_key):
    """Petrolfarbenes Kopfband mit Titel und Unterzeile."""
    paint(ws, c1, 2, c2, 4, PETROL)
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 16
    l1, l2 = get_column_letter(c1), get_column_letter(c2)
    # Rechts vom Titel bleibt nur dort Platz, wo der Sprachumschalter sitzt.
    end = get_column_letter(c2 - 4 if c2 - 4 > c1 + 1 else c2)
    ws.merge_cells("%s3:%s3" % (l1, end))
    ws.merge_cells("%s4:%s4" % (l1, end))
    put(ws, "%s3" % l1, T(title_key),
        font=Font(name=FONT, size=18, bold=True, color="FFFFFF"),
        align=Alignment(horizontal="left", vertical="center", indent=1))
    put(ws, "%s4" % l1, T(sub_key),
        font=Font(name=FONT, size=9.5, color="A9C4D0"),
        align=Alignment(horizontal="left", vertical="top", indent=1))
    return l1, l2


def section(ws, coord, key, span_to=None):
    """Abschnittsueberschrift im Corporate-Stil."""
    if span_to:
        ws.merge_cells("%s:%s" % (coord, span_to))
    put(ws, coord, T(key),
        font=Font(name=FONT, size=11, bold=True, color=PETROL),
        align=Alignment(horizontal="left", vertical="center"))


def dxf_rule(formula, fill=None, color=None, bold=False, numfmt=None, fmt_id=200):
    kw = {}
    if fill:
        kw["fill"] = PatternFill(bgColor=fill)
    if color or bold:
        kw["font"] = Font(color=color, bold=bold or None)
    if numfmt:
        kw["numFmt"] = NumberFormat(numFmtId=fmt_id, formatCode=numfmt)
    return Rule(type="expression", formula=[formula], dxf=DifferentialStyle(**kw))


def en_format(ws, ref, code, fmt_id):
    """Schaltet das Zahlenformat eines Bereichs auf die englische Variante um."""
    ws.conditional_formatting.add(ref, dxf_rule("LANG=2", numfmt=code, fmt_id=fmt_id))


# ---------------------------------------------------------------------------
# 5. AUFBAU DER ARBEITSMAPPE
# ---------------------------------------------------------------------------

def build(path):
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    ws_dash = wb.active
    ws_dash.title = SH_DASH
    ws_inv  = wb.create_sheet(SH_INV)
    ws_cost = wb.create_sheet(SH_COST)
    ws_scn  = wb.create_sheet(SH_SCN)
    ws_cfg  = wb.create_sheet(SH_CFG)

    ws_dash.sheet_properties.tabColor = PETROL
    ws_inv.sheet_properties.tabColor  = TEAL
    ws_cost.sheet_properties.tabColor = TEAL
    ws_scn.sheet_properties.tabColor  = SAGE
    ws_cfg.sheet_properties.tabColor  = MUTED

    build_config(ws_cfg)
    build_invoices(ws_inv)
    build_costs(ws_cost)
    build_scenario(ws_scn)
    build_dashboard(ws_dash)
    build_engine(ws_cfg)
    define_names(wb)

    ws_cfg.sheet_state = "hidden"
    wb.active = 0
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 5a. CONFIG-BLATT: Uebersetzungen, Listen, Rechen-Engine
# ---------------------------------------------------------------------------

def build_config(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    for col in "EFGHIJ":
        ws.column_dimensions[col].width = 18

    put(ws, "A1", "CONFIG — Steuerdaten. Bitte nicht verändern. / Control data. Please do not edit.",
        font=Font(name=FONT, size=11, bold=True, color=RED_FG))

    # Sprach- und USt-Schalter (Ziel der benannten Bereiche LANG / VATON)
    put(ws, "D2", "LANG", font=Font(name=FONT, size=9, color=MUTED))
    put(ws, "E2", '=IF({d}!$N$3="English",2,1)'.format(d=SH_DASH),
        font=Font(name=FONT, size=9, bold=True))
    put(ws, "D3", "VATON", font=Font(name=FONT, size=9, color=MUTED))
    put(ws, "E3", '=IF(COUNTIF($F$8:$G$8,{d}!$N$37)>0,1,0)'.format(d=SH_DASH),
        font=Font(name=FONT, size=9, bold=True))

    # Uebersetzungstabelle
    for i, txt in enumerate(("Key", "Deutsch", "English")):
        put(ws, "%s5" % "ABC"[i], txt,
            font=Font(name=FONT, size=9, bold=True, color="FFFFFF"),
            fill=PETROL)
    for i, (key, de, en) in enumerate(TRANSLATIONS):
        r = TR_FIRST + i
        put(ws, "A%d" % r, key, font=Font(name=FONT, size=9, color=MUTED))
        put(ws, "B%d" % r, de, font=Font(name=FONT, size=9))
        put(ws, "C%d" % r, en, font=Font(name=FONT, size=9))

    # Statuswerte: Spalte F deutsch, G englisch, H sprachabhaengige Anzeige.
    # Die Auswertung prueft immer beide Sprachen, damit ein Sprachwechsel
    # bereits erfasste Daten nicht entwertet.
    put(ws, "F4", "DE",   font=Font(name=FONT, size=9, bold=True, color=MUTED))
    put(ws, "G4", "EN",   font=Font(name=FONT, size=9, bold=True, color=MUTED))
    put(ws, "H4", "aktiv/active", font=Font(name=FONT, size=9, bold=True, color=MUTED))
    pairs = [(5, "offen", "open"), (6, "bezahlt", "paid"),
             (8, "Ja", "Yes"),     (9, "Nein", "No")]
    for r, de, en in pairs:
        put(ws, "F%d" % r, de, font=Font(name=FONT, size=9))
        put(ws, "G%d" % r, en, font=Font(name=FONT, size=9))
        put(ws, "H%d" % r, "=INDEX(F%d:G%d,LANG)" % (r, r), font=Font(name=FONT, size=9))

    # Eindeutige Kundenliste fuer das Szenario-Dropdown (ohne Matrixformel,
    # damit sie auch in Google Sheets ohne Sondertasten funktioniert).
    put(ws, "J4", "Kunden / Customers", font=Font(name=FONT, size=9, bold=True, color=MUTED))
    for r in range(5, 40):
        ws["J%d" % r] = (
            '=IFERROR(INDEX({inv}!$C$9:$C$108,'
            'MATCH(0,INDEX(COUNTIF($J$4:J{prev},{inv}!$C$9:$C$108)'
            '+({inv}!$C$9:$C$108=""),0),0)),"")'
        ).format(inv=QINV, prev=r - 1)
        ws["J%d" % r].font = Font(name=FONT, size=9)


def build_engine(ws):
    """12-Wochen-Rechenkern in den Spalten L..Y plus Fixkostenmatrix AA..AM."""
    hdr = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    for col, label in zip("LMNOPQRSTUVWXY", [
            "Wo", "Start", "Ende", "Ein Basis", "Ein Szenario", "Aus fix",
            "Aus variabel", "Aus USt", "Aus gesamt", "Saldo Basis",
            "Saldo Szenario", "Ampel", "Reserve", "Label"]):
        put(ws, "%s4" % col, label, font=hdr, fill=PETROL_SOFT)
        ws.column_dimensions[col].width = 15

    # Serientitel fuer das Diagramm liegen sprachabhaengig in U4/V4/X4.
    put(ws, "U4", T("ser.base"), font=hdr, fill=PETROL_SOFT)
    put(ws, "V4", T("ser.scn"),  font=hdr, fill=PETROL_SOFT)
    put(ws, "X4", T("ser.min"),  font=hdr, fill=PETROL_SOFT)

    for i in range(12):
        r = 5 + i
        mcol = get_column_letter(28 + i)          # AB .. AM
        ws["L%d" % r] = i + 1
        ws["M%d" % r] = "=START" if i == 0 else "=M%d+7" % (r - 1)
        ws["N%d" % r] = "=M%d+6" % r
        ws["M%d" % r].number_format = FMT_DATE_DE
        ws["N%d" % r].number_format = FMT_DATE_DE

        # Einzahlungen: Basis ueber das reale Faelligkeitsdatum,
        # Szenario ueber das im Szenario verschobene Datum.
        ws["O%d" % r] = ('=SUMIFS({i}!$Q$9:$Q$108,{i}!$R$9:$R$108,">="&M{r},'
                         '{i}!$R$9:$R$108,"<="&N{r})').format(i=QINV, r=r)
        ws["P%d" % r] = ('=SUMIFS({i}!$S$9:$S$108,{i}!$L$9:$L$108,">="&M{r},'
                         '{i}!$L$9:$L$108,"<="&N{r})').format(i=QINV, r=r)

        # Auszahlungen
        ws["Q%d" % r] = "=%s$25" % mcol
        ws["R%d" % r] = ('=SUMIFS({c}!$D$34:$D$63,{c}!$E$34:$E$63,">="&M{r},'
                         '{c}!$E$34:$E$63,"<="&N{r})').format(c=QCOST, r=r)
        # Umsatzsteuer-Zahllast: am VATDAY des Monats fuer die im Vormonat
        # gestellten Rechnungen (Soll-Versteuerung). Bei VATON=0 immer 0.
        ws["S%d" % r] = (
            '=IF(VATON=0,0,'
            'IF(AND(DATE(YEAR(M{r}),MONTH(M{r}),VATDAY)>=M{r},'
            'DATE(YEAR(M{r}),MONTH(M{r}),VATDAY)<=N{r}),'
            'SUMIFS({i}!$P$9:$P$108,{i}!$H$9:$H$108,">="&DATE(YEAR(M{r}),MONTH(M{r})-1,1),'
            '{i}!$H$9:$H$108,"<="&DATE(YEAR(M{r}),MONTH(M{r}),1)-1),0)'
            '+IF(AND(MONTH(M{r})<>MONTH(N{r}),'
            'DATE(YEAR(N{r}),MONTH(N{r}),VATDAY)>=M{r},'
            'DATE(YEAR(N{r}),MONTH(N{r}),VATDAY)<=N{r}),'
            'SUMIFS({i}!$P$9:$P$108,{i}!$H$9:$H$108,">="&DATE(YEAR(N{r}),MONTH(N{r})-1,1),'
            '{i}!$H$9:$H$108,"<="&DATE(YEAR(N{r}),MONTH(N{r}),1)-1),0))'
        ).format(i=QINV, r=r)
        ws["T%d" % r] = "=Q{r}+R{r}+S{r}".format(r=r)

        ws["U%d" % r] = ("=OPENBAL+O%d-T%d" % (r, r)) if i == 0 else \
                        ("=U%d+O%d-T%d" % (r - 1, r, r))
        ws["V%d" % r] = ("=OPENBAL+P%d-T%d" % (r, r)) if i == 0 else \
                        ("=V%d+P%d-T%d" % (r - 1, r, r))
        ws["W%d" % r] = ("=IF(V{r}<MINCASH,2,IF(V{r}<MINCASH*YELLOWF,1,0))".format(r=r))
        ws["X%d" % r] = "=MINCASH"
        ws["Y%d" % r] = '={t}&" "&ISOWEEKNUM(M{r})'.format(t=Traw("lbl.week"), r=r)
        for col in "OPQRSTUVX":
            ws["%s%d" % (col, r)].number_format = FMT_EUR0_DE

    # ---- Fixkostenmatrix: je Kostenzeile x je Prognosewoche ---------------
    put(ws, "AA4", "Tag", font=Font(name=FONT, size=9, bold=True, color=MUTED))
    for i in range(12):
        mcol = get_column_letter(28 + i)
        ws["%s3" % mcol] = "=M%d" % (5 + i)
        ws["%s4" % mcol] = "=N%d" % (5 + i)
        ws["%s3" % mcol].number_format = FMT_DATE_DE
        ws["%s4" % mcol].number_format = FMT_DATE_DE

    for j in range(20):                            # 20 Fixkostenzeilen
        r = 5 + j
        crow = 9 + j                               # Zeile im Kostenblatt
        ws["AA%d" % r] = ('=IF({c}!$E${cr}="",1,MIN(28,MAX(1,{c}!$E${cr})))'
                          .format(c=QCOST, cr=crow))
        for i in range(12):
            mcol = get_column_letter(28 + i)
            ws["%s%d" % (mcol, r)] = (
                '=IF(N({c}!$D${cr})=0,0,'
                'IF(OR(AND(DATE(YEAR({m}$3),MONTH({m}$3),$AA{r})>={m}$3,'
                'DATE(YEAR({m}$3),MONTH({m}$3),$AA{r})<={m}$4),'
                'AND(DATE(YEAR({m}$4),MONTH({m}$4),$AA{r})>={m}$3,'
                'DATE(YEAR({m}$4),MONTH({m}$4),$AA{r})<={m}$4)),'
                '{c}!$D${cr},0))'
            ).format(c=QCOST, cr=crow, m=mcol, r=r)
    for i in range(12):
        mcol = get_column_letter(28 + i)
        ws["%s25" % mcol] = "=SUM({m}5:{m}24)".format(m=mcol)


# ---------------------------------------------------------------------------
# 5b. RECHNUNGEN
# ---------------------------------------------------------------------------

INV_FIRST, INV_LAST = 9, 108


def build_invoices(ws):
    canvas(ws, 20, INV_LAST + 6)
    widths = {"A": 2, "B": 5.5, "C": 26, "D": 14, "E": 14, "F": 9.5, "G": 14,
              "H": 15, "I": 16, "J": 14, "K": 13, "L": 18, "M": 18, "N": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in "OPQRS":
        ws.column_dimensions[col].hidden = True
        ws.column_dimensions[col].width = 12

    header_band(ws, 2, 13, "app.title", "inv.title")

    ws.row_dimensions[5].height = 8
    section(ws, "B6", "inv.title", "F6")
    ws.merge_cells("B7:M7")
    put(ws, "B7", T("inv.hint"),
        font=Font(name=FONT, size=9, color=MUTED),
        align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 16

    # Tabellenkopf
    heads = [("B", "inv.h.no"), ("C", "inv.h.customer"), ("D", "inv.h.invno"),
             ("E", "inv.h.amount"), ("F", "inv.h.vat"), ("G", "inv.h.gross"),
             ("H", "inv.h.date"), ("I", "inv.h.terms"), ("J", "inv.h.due"),
             ("K", "inv.h.status"), ("L", "inv.h.expdate"), ("M", "inv.h.expamt")]
    ws.row_dimensions[8].height = 30
    for col, key in heads:
        put(ws, "%s8" % col, T(key),
            font=Font(name=FONT, size=9.5, bold=True, color="FFFFFF"),
            fill=PETROL,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True))

    input_cols = "CDEFHIK"
    calc_cols  = "GJLM"

    for i in range(INV_LAST - INV_FIRST + 1):
        r = INV_FIRST + i
        ws.row_dimensions[r].height = 17
        zebra = ZEBRA if i % 2 else CARD

        put(ws, "B%d" % r, i + 1,
            font=Font(name=FONT, size=9, color=MUTED), fill=zebra,
            align=Alignment(horizontal="center", vertical="center"))
        for col in input_cols:
            put(ws, "%s%d" % (col, r), None, font=Font(name=FONT, size=10, color=TEXT),
                fill=zebra, align=Alignment(horizontal="left", vertical="center", indent=1))
        for col in calc_cols:
            put(ws, "%s%d" % (col, r), None,
                font=Font(name=FONT, size=10, color=MUTED, italic=True),
                fill=CALC_BG, align=Alignment(horizontal="right", vertical="center", indent=1))

        # Formeln der berechneten Spalten
        ws["G%d" % r] = '=IF($E{r}="","",$E{r}*(1+IF(VATON=1,N($F{r}),0)))'.format(r=r)
        ws["J%d" % r] = '=IF($H{r}="","",$H{r}+N($I{r}))'.format(r=r)
        ws["L%d" % r] = ('=IF($O{r}=0,"",MAX($J{r},START)+DELAY_ALL'
                         '+IF($C{r}=SCN_CUSTOMER,DELAY_CUSTOMER,0))').format(r=r)
        ws["M%d" % r] = '=IF($O{r}=0,"",$S{r})'.format(r=r)

        # Ausgeblendete Hilfsspalten
        ws["O%d" % r] = ('=IF($K{r}="",0,IF(COUNTIF({c}!$F$5:$G$5,$K{r})>0,1,0))'
                         .format(r=r, c=SH_CFG))
        ws["P%d" % r] = '=IF($E{r}="",0,IF(VATON=1,$E{r}*N($F{r}),0))'.format(r=r)
        ws["Q%d" % r] = '=IF($O{r}=0,0,N($G{r}))'.format(r=r)
        ws["R%d" % r] = '=IF($O{r}=0,"",MAX($J{r},START))'.format(r=r)
        ws["S%d" % r] = ('=IF($O{r}=0,0,N($G{r})*(1-IF($C{r}=SCN_CUSTOMER,LOSS_RATE,0)))'
                         .format(r=r))

        for col in "EGM":
            ws["%s%d" % (col, r)].number_format = FMT_EUR_DE
        for col in "HJLR":
            ws["%s%d" % (col, r)].number_format = FMT_DATE_DE
        ws["F%d" % r].number_format = FMT_PCT
        ws["I%d" % r].alignment = Alignment(horizontal="center", vertical="center")
        ws["F%d" % r].alignment = Alignment(horizontal="center", vertical="center")
        ws["E%d" % r].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws["K%d" % r].alignment = Alignment(horizontal="center", vertical="center")

    # Beispieldaten
    for i, (cust, no, amount, d_inv, terms, paid) in enumerate(INVOICES):
        r = INV_FIRST + i
        ws["C%d" % r] = cust
        ws["D%d" % r] = no
        ws["E%d" % r] = amount
        # Der USt-Satz erscheint nur, wenn der Umsatzsteuer-Schalter aktiv ist.
        # So bleibt die Demo im §19-Modus widerspruchsfrei (brutto = netto) und
        # zeigt beim Umlegen des Schalters sofort die volle USt-Logik.
        ws["F%d" % r] = '=IF(VATON=1,0.19,"")'
        ws["H%d" % r] = "=TODAY()%+d" % d_inv
        ws["I%d" % r] = terms
        ws["K%d" % r] = "=%s!$H$%d" % (SH_CFG, 6 if paid else 5)

    # Summenzeile
    total = INV_LAST + 1
    ws.row_dimensions[total].height = 24
    box(ws, 2, total, 13, total, fill=PETROL)
    ws.merge_cells("B%d:F%d" % (total, total))
    put(ws, "B%d" % total, T("inv.total"),
        font=Font(name=FONT, size=10, bold=True, color="FFFFFF"),
        align=Alignment(horizontal="left", vertical="center", indent=1))
    put(ws, "G%d" % total, "=SUM($Q$%d:$Q$%d)" % (INV_FIRST, INV_LAST),
        font=Font(name=FONT, size=11, bold=True, color="FFFFFF"),
        align=Alignment(horizontal="right", vertical="center", indent=1),
        fmt=FMT_EUR_DE)
    ws.merge_cells("L%d:M%d" % (total, total))
    put(ws, "L%d" % total, "=SUM($S$%d:$S$%d)" % (INV_FIRST, INV_LAST),
        font=Font(name=FONT, size=11, bold=True, color="FFFFFF"),
        align=Alignment(horizontal="right", vertical="center", indent=1),
        fmt=FMT_EUR_DE)

    note = total + 2
    ws.merge_cells("B%d:M%d" % (note, note))
    put(ws, "B%d" % note, T("inv.scnnote"),
        font=Font(name=FONT, size=9, italic=True, color=MUTED),
        align=Alignment(horizontal="left", vertical="center"))

    # Datenpruefung
    dv_status = DataValidation(type="list", formula1="%s!$H$5:$H$6" % SH_CFG,
                               allow_blank=True, showDropDown=False)
    dv_terms  = DataValidation(type="whole", operator="between",
                               formula1="0", formula2="365", allow_blank=True)
    dv_vat    = DataValidation(type="decimal", operator="between",
                               formula1="0", formula2="1", allow_blank=True)
    for dv in (dv_status, dv_terms, dv_vat):
        ws.add_data_validation(dv)
    dv_status.add("K%d:K%d" % (INV_FIRST, INV_LAST))
    dv_terms.add("I%d:I%d" % (INV_FIRST, INV_LAST))
    dv_vat.add("F%d:F%d" % (INV_FIRST, INV_LAST))

    # Ueberfaellige offene Rechnungen hervorheben
    rng = "B%d:M%d" % (INV_FIRST, INV_LAST)
    ws.conditional_formatting.add(
        rng, dxf_rule('AND($O{f}=1,$J{f}<>"",$J{f}<TODAY())'.format(f=INV_FIRST),
                      fill=RED_BG, color=RED_FG))

    # Englische Zahlen- und Datumsformate
    en_format(ws, "E%d:E%d" % (INV_FIRST, INV_LAST + 1), FMT_EUR_EN, 201)
    en_format(ws, "G%d:G%d" % (INV_FIRST, INV_LAST + 1), FMT_EUR_EN, 202)
    en_format(ws, "M%d:M%d" % (INV_FIRST, INV_LAST + 1), FMT_EUR_EN, 203)
    en_format(ws, "H%d:H%d" % (INV_FIRST, INV_LAST), FMT_DATE_EN, 210)
    en_format(ws, "J%d:J%d" % (INV_FIRST, INV_LAST), FMT_DATE_EN, 211)
    en_format(ws, "L%d:L%d" % (INV_FIRST, INV_LAST), FMT_DATE_EN, 212)

    ws.freeze_panes = "B9"


# ---------------------------------------------------------------------------
# 5c. KOSTEN
# ---------------------------------------------------------------------------

FIX_FIRST, FIX_LAST = 9, 28
VAR_FIRST, VAR_LAST = 34, 63


def build_costs(ws):
    canvas(ws, 8, VAR_LAST + 6)
    for col, w in {"A": 2, "B": 38, "C": 24, "D": 17, "E": 20, "F": 2}.items():
        ws.column_dimensions[col].width = w

    header_band(ws, 2, 5, "app.title", "cost.title")

    def table(title_key, hint_key, hrow, first, last, headers, money_fmt, last_fmt):
        ws.row_dimensions[hrow - 2].height = 20
        section(ws, "B%d" % (hrow - 2), title_key, "E%d" % (hrow - 2))
        ws.merge_cells("B%d:E%d" % (hrow - 1, hrow - 1))
        put(ws, "B%d" % (hrow - 1), T(hint_key),
            font=Font(name=FONT, size=9, color=MUTED),
            align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[hrow - 1].height = 16
        ws.row_dimensions[hrow].height = 24
        for col, key in headers:
            put(ws, "%s%d" % (col, hrow), T(key),
                font=Font(name=FONT, size=9.5, bold=True, color="FFFFFF"),
                fill=PETROL,
                align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        for i in range(last - first + 1):
            r = first + i
            ws.row_dimensions[r].height = 17
            zebra = ZEBRA if i % 2 else CARD
            for col in "BCDE":
                put(ws, "%s%d" % (col, r), None,
                    font=Font(name=FONT, size=10, color=TEXT), fill=zebra,
                    align=Alignment(horizontal="left", vertical="center", indent=1))
            ws["D%d" % r].number_format = money_fmt
            ws["D%d" % r].alignment = Alignment(horizontal="right", vertical="center", indent=1)
            ws["E%d" % r].number_format = last_fmt
            ws["E%d" % r].alignment = Alignment(horizontal="center", vertical="center")
        # Summenzeile
        s = last + 1
        ws.row_dimensions[s].height = 24
        box(ws, 2, s, 5, s, fill=PETROL)
        ws.merge_cells("B%d:C%d" % (s, s))
        put(ws, "B%d" % s, T("cost.sum.fixed" if first == FIX_FIRST else "cost.sum.var"),
            font=Font(name=FONT, size=10, bold=True, color="FFFFFF"),
            align=Alignment(horizontal="left", vertical="center", indent=1))
        put(ws, "D%d" % s, "=SUM(D%d:D%d)" % (first, last),
            font=Font(name=FONT, size=11, bold=True, color="FFFFFF"),
            align=Alignment(horizontal="right", vertical="center", indent=1),
            fmt=money_fmt)

    table("cost.fixed", "cost.fixed.hint", 8, FIX_FIRST, FIX_LAST,
          [("B", "cost.h.name"), ("C", "cost.h.cat"),
           ("D", "cost.h.month"), ("E", "cost.h.day")],
          FMT_EUR_DE, "0")
    table("cost.var", "cost.var.hint", 33, VAR_FIRST, VAR_LAST,
          [("B", "cost.h.name"), ("C", "cost.h.cat"),
           ("D", "cost.h.amount"), ("E", "cost.h.date")],
          FMT_EUR_DE, FMT_DATE_DE)

    for i, (name, cat, amount, day) in enumerate(FIXED_COSTS):
        r = FIX_FIRST + i
        ws["B%d" % r], ws["C%d" % r] = name, cat
        ws["D%d" % r], ws["E%d" % r] = amount, day
    for i, (name, cat, amount, offset) in enumerate(VAR_COSTS):
        r = VAR_FIRST + i
        ws["B%d" % r], ws["C%d" % r] = name, cat
        ws["D%d" % r] = amount
        ws["E%d" % r] = "=TODAY()%+d" % offset

    dv_day = DataValidation(type="whole", operator="between",
                            formula1="1", formula2="28", allow_blank=True)
    ws.add_data_validation(dv_day)
    dv_day.add("E%d:E%d" % (FIX_FIRST, FIX_LAST))

    en_format(ws, "D%d:D%d" % (FIX_FIRST, FIX_LAST + 1), FMT_EUR_EN, 204)
    en_format(ws, "D%d:D%d" % (VAR_FIRST, VAR_LAST + 1), FMT_EUR_EN, 205)
    en_format(ws, "E%d:E%d" % (VAR_FIRST, VAR_LAST), FMT_DATE_EN, 213)

    ws.freeze_panes = "B9"


# ---------------------------------------------------------------------------
# 5d. SZENARIO
# ---------------------------------------------------------------------------

def grid_columns(ws):
    """Einheitliches Spaltenraster fuer Dashboard und Szenario: 4 Kacheln."""
    ws.column_dimensions["A"].width = 2
    for col in "BCDFGHJKLNOP":
        ws.column_dimensions[col].width = 11
    for col in "EIM":
        ws.column_dimensions[col].width = 2
    ws.column_dimensions["Q"].width = 2


def kpi_card(ws, c1, r_top, label_key, value_formula, sub_formula,
             value_fmt=FMT_EUR0_DE, value_color=PETROL, fill=CARD, accent=TEAL):
    """Kennzahlen-Kachel ueber drei Spalten und vier Zeilen.

    Die oberste Zeile der Kachel bleibt schmal und dient als farbige
    Akzentkante - das gibt der Flaeche den Charakter einer BI-Karte.
    """
    c2 = c1 + 2
    box(ws, c1, r_top, c2, r_top + 3, fill=fill)
    paint(ws, c1, r_top, c2, r_top, accent)
    l1, l2 = get_column_letter(c1), get_column_letter(c2)
    ws.merge_cells("%s%d:%s%d" % (l1, r_top + 1, l2, r_top + 1))
    ws.merge_cells("%s%d:%s%d" % (l1, r_top + 2, l2, r_top + 2))
    put(ws, "%s%d" % (l1, r_top + 1), T(label_key),
        font=Font(name=FONT, size=8.5, bold=True, color=MUTED),
        align=Alignment(horizontal="left", vertical="center", indent=2))
    put(ws, "%s%d" % (l1, r_top + 2), value_formula,
        font=Font(name=FONT, size=17, bold=True, color=value_color),
        align=Alignment(horizontal="left", vertical="center", indent=1),
        fmt=value_fmt)
    if sub_formula is not None:
        ws.merge_cells("%s%d:%s%d" % (l1, r_top + 3, l2, r_top + 3))
        put(ws, "%s%d" % (l1, r_top + 3), sub_formula,
            font=Font(name=FONT, size=8.5, color=MUTED),
            align=Alignment(horizontal="left", vertical="top", indent=2))
    return "%s%d:%s%d" % (l1, r_top, l2, r_top + 3)


def control_row(ws, row, label_key, hint_key, value, fmt, suffix_key=None):
    """Eine Eingabezeile im Reglerblock des Szenario-Tabs."""
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 15
    ws.merge_cells("C%d:H%d" % (row, row))
    put(ws, "C%d" % row, T(label_key),
        font=Font(name=FONT, size=11, bold=True, color=TEXT),
        align=Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("C%d:H%d" % (row + 1, row + 1))
    put(ws, "C%d" % (row + 1), T(hint_key),
        font=Font(name=FONT, size=8.5, color=MUTED),
        align=Alignment(horizontal="left", vertical="top"))
    ws.merge_cells("J%d:L%d" % (row, row))
    box(ws, 10, row, 12, row, fill=INPUT_BG, border=INPUT_BD)
    put(ws, "J%d" % row, value,
        font=Font(name=FONT, size=13, bold=True, color=PETROL),
        align=Alignment(horizontal="center", vertical="center"), fmt=fmt)
    if suffix_key:
        put(ws, "N%d" % row, T(suffix_key),
            font=Font(name=FONT, size=9, color=MUTED),
            align=Alignment(horizontal="left", vertical="center"))


def build_scenario(ws):
    canvas(ws, 17, 46)
    grid_columns(ws)
    header_band(ws, 2, 16, "app.title", "scn.title")

    ws.row_dimensions[5].height = 8
    section(ws, "B6", "scn.controls", "H6")
    ws.merge_cells("B7:P7")
    put(ws, "B7", T("scn.hint"),
        font=Font(name=FONT, size=9, color=MUTED),
        align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 6

    box(ws, 2, 9, 16, 20, fill=CARD)
    control_row(ws, 10, "scn.delay_all",   "scn.delay_all.h",   SCN_DELAY_ALL,  "0", "scn.days")
    control_row(ws, 13, "scn.customer",    "scn.customer.h",    SCN_CUSTOMER,   "@")
    control_row(ws, 16, "scn.delay_cust",  "scn.delay_cust.h",  SCN_DELAY_CUST, "0", "scn.days")
    control_row(ws, 19, "scn.loss",        "scn.loss.h",        SCN_LOSS,       FMT_PCT)
    ws["J13"].font = Font(name=FONT, size=11, bold=True, color=PETROL)

    ws.merge_cells("B21:P21")
    put(ws, "B21", T("scn.reset"),
        font=Font(name=FONT, size=8.5, italic=True, color=MUTED),
        align=Alignment(horizontal="left", vertical="center"))

    dv_cust = DataValidation(type="list", formula1="%s!$J$5:$J$39" % SH_CFG,
                             allow_blank=True, showDropDown=False)
    dv_days = DataValidation(type="whole", operator="between",
                             formula1="0", formula2="365", allow_blank=True)
    dv_loss = DataValidation(type="decimal", operator="between",
                             formula1="0", formula2="1", allow_blank=True)
    for dv in (dv_cust, dv_days, dv_loss):
        ws.add_data_validation(dv)
    dv_cust.add("J13")
    dv_days.add("J10")
    dv_days.add("J16")
    dv_loss.add("J19")

    # --- Wirkungskennzahlen ------------------------------------------------
    ws.row_dimensions[22].height = 8
    section(ws, "B23", "scn.impact", "P23")
    ws.row_dimensions[23].height = 20
    ws.row_dimensions[24].height = 6
    for r, h in ((25, 6), (26, 15), (27, 26), (28, 14)):
        ws.row_dimensions[r].height = h

    kpi_card(ws, 2,  25, "scn.low.base",
             "=MIN(%s!$U$5:$U$16)" % SH_CFG, None)
    kpi_card(ws, 6,  25, "scn.low.scn",
             "=MIN(%s!$V$5:$V$16)" % SH_CFG, None, accent=SAGE)
    kpi_card(ws, 10, 25, "scn.gap",
             "=MIN({c}!$V$5:$V$16)-MIN({c}!$U$5:$U$16)".format(c=SH_CFG), None,
             value_color=RED_FG, accent=RED_FG)
    kpi_card(ws, 14, 25, "scn.firstwarn",
             ('=IF(MAX({c}!$W$5:$W$16)=0,{n},'
              'INDEX({c}!$Y$5:$Y$16,MATCH(MAX({c}!$W$5:$W$16),{c}!$W$5:$W$16,0)))'
              ).format(c=SH_CFG, n=Traw("scn.none")),
             None, value_fmt="@")

    # --- Vergleichstabelle und Diagramm ------------------------------------
    thead = 31
    ws.row_dimensions[29].height = 10
    ws.row_dimensions[30].height = 20
    section(ws, "B30", "sec.table", "F30")
    section(ws, "H30", "scn.chart", "P30")
    ws.row_dimensions[thead].height = 22
    heads = [("B", "th.week"), ("C", "th.base"), ("D", "th.scn"), ("F", "scn.h.diff")]
    for col, key in heads:
        put(ws, "%s%d" % (col, thead), T(key),
            font=Font(name=FONT, size=9, bold=True, color="FFFFFF"), fill=PETROL,
            align=Alignment(horizontal="center", vertical="center"))
    put(ws, "E%d" % thead, None, fill=PETROL)
    for i in range(12):
        r = thead + 1 + i
        ws.row_dimensions[r].height = 17
        zebra = ZEBRA if i % 2 else CARD
        box(ws, 2, r, 6, r, fill=zebra)
        put(ws, "B%d" % r, "=%s!$Y$%d" % (SH_CFG, 5 + i),
            font=Font(name=FONT, size=9.5, bold=True, color=PETROL),
            align=Alignment(horizontal="center", vertical="center"))
        put(ws, "C%d" % r, "=%s!$U$%d" % (SH_CFG, 5 + i),
            font=Font(name=FONT, size=10, color=TEXT), fmt=FMT_EUR0_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))
        ws.merge_cells("D%d:E%d" % (r, r))
        put(ws, "D%d" % r, "=%s!$V$%d" % (SH_CFG, 5 + i),
            font=Font(name=FONT, size=10, bold=True, color=TEXT), fmt=FMT_EUR0_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))
        put(ws, "F%d" % r, "={c}!$V${x}-{c}!$U${x}".format(c=SH_CFG, x=5 + i),
            font=Font(name=FONT, size=10, color=MUTED), fmt=FMT_EUR0Z_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))

    first, last = thead + 1, thead + 12
    traffic_cf(ws, "B%d:B%d" % (first, last), first, chip=True)
    traffic_cf(ws, "D%d:D%d" % (first, last), first, chip=False)
    en_format(ws, "C%d:E%d" % (first, last), FMT_EUR0_EN, 206)
    en_format(ws, "F%d:F%d" % (first, last), FMT_EUR0Z_EN, 217)

    box(ws, 8, thead, 16, last, fill=CARD)
    chart = build_chart(width=15.0, height=7.0)
    ws.add_chart(chart, "H%d" % thead)

    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def traffic_cf(ws, ref, first_row, chip):
    """Ampel-Formatierung ueber den Ampelcode der Rechen-Engine."""
    idx = '=INDEX({c}!$W$5:$W$16,ROW()-{o})'.format(c=SH_CFG, o=first_row - 1)
    combos = [(2, RED_BG, RED_FG), (1, AMBER_BG, AMBER_FG), (0, GREEN_BG, GREEN_FG)]
    for code, bg, fg in combos:
        ws.conditional_formatting.add(
            ref, dxf_rule("%s=%d" % (idx, code),
                          fill=bg if chip else None, color=fg, bold=True))


# ---------------------------------------------------------------------------
# 5e. DASHBOARD
# ---------------------------------------------------------------------------

TBL_HEAD  = 32
TBL_FIRST = 33
TBL_LAST  = 44


def build_dashboard(ws):
    canvas(ws, 17, 56)
    grid_columns(ws)

    # ---- Kopfband mit Sprachumschalter -----------------------------------
    header_band(ws, 2, 16, "app.title", "app.sub")
    put(ws, "N2", T("ui.language"),
        font=Font(name=FONT, size=8, color="A9C4D0"),
        align=Alignment(horizontal="center", vertical="bottom"))
    ws.merge_cells("N2:P2")
    ws.merge_cells("N3:P3")
    box(ws, 14, 3, 16, 3, fill="FFFFFF", border="6E8C9C")
    put(ws, "N3", "Deutsch",
        font=Font(name=FONT, size=11, bold=True, color=PETROL),
        align=Alignment(horizontal="center", vertical="center"))
    dv_lang = DataValidation(type="list", formula1='"Deutsch,English"',
                             allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv_lang)
    dv_lang.add("N3")

    # ---- Kennzahlen-Kacheln ----------------------------------------------
    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[7].height = 15
    ws.row_dimensions[8].height = 26
    ws.row_dimensions[9].height = 14
    ws.row_dimensions[10].height = 6

    kpi_card(ws, 2, 6, "kpi.balance", "=OPENBAL", T("kpi.balance.sub"))
    kpi_card(ws, 6, 6, "kpi.recv",
             "=SUM({i}!$Q$9:$Q$108)".format(i=QINV),
             '=COUNTIF({i}!$O$9:$O$108,1)&" "&{s}'.format(i=QINV, s=Traw("kpi.recv.sub")),
             value_color=TEAL)
    kpi_card(ws, 10, 6, "kpi.low",
             "=MIN(%s!$V$5:$V$16)" % SH_CFG, T("kpi.low.sub"))
    kpi_card(
        ws, 14, 6, "kpi.status",
        ('=IF(MAX({c}!$W$5:$W$16)=2,{r},IF(MAX({c}!$W$5:$W$16)=1,{a},{g}))'
         ).format(c=SH_CFG, r=Traw("status.red"), a=Traw("status.amber"),
                  g=Traw("status.green")),
        T("kpi.status.sub"), value_fmt="@", value_color=GREEN_FG, accent=GREEN_FG)

    # Statuskachel: die Akzentkante traegt die Ampelfarbe, der Wert die Schrift-
    # farbe. Der Kachelkoerper bleibt weiss, damit alle vier Kacheln gleich wirken.
    for code, bg, fg in ((2, RED_FG, RED_FG), (1, AMBER_FG, AMBER_FG), (0, GREEN_FG, GREEN_FG)):
        rule = "MAX({c}!$W$5:$W$16)={v}".format(c=SH_CFG, v=code)
        ws.conditional_formatting.add("N6:P6", dxf_rule(rule, fill=bg))
        ws.conditional_formatting.add("N8:P8", dxf_rule(rule, color=fg, bold=True))
    # Tiefststand-Kachel: Wert rot, sobald er unter der Reserve liegt
    ws.conditional_formatting.add(
        "J8:L8", dxf_rule("MIN({c}!$V$5:$V$16)<MINCASH".format(c=SH_CFG),
                          color=RED_FG, bold=True))

    # ---- Warnbanner -------------------------------------------------------
    ws.row_dimensions[11].height = 6
    ws.row_dimensions[12].height = 26
    box(ws, 2, 12, 16, 12, fill=GREEN_BG, border=None)
    ws.merge_cells("B12:P12")
    put(ws, "B12",
        ('=IF(MAX({c}!$W$5:$W$16)=0,{g},'
         'IF(MAX({c}!$W$5:$W$16)=2,{r},{a})&" "'
         '&INDEX({c}!$Y$5:$Y$16,MATCH(MAX({c}!$W$5:$W$16),{c}!$W$5:$W$16,0))&".")'
         ).format(c=SH_CFG, g=Traw("alert.green"), r=Traw("alert.red"),
                  a=Traw("alert.amber")),
        font=Font(name=FONT, size=10.5, bold=True, color=GREEN_FG),
        align=Alignment(horizontal="left", vertical="center", indent=2))
    for code, bg, fg in ((2, RED_BG, RED_FG), (1, AMBER_BG, AMBER_FG), (0, GREEN_BG, GREEN_FG)):
        ws.conditional_formatting.add(
            "B12:P12",
            dxf_rule("MAX({c}!$W$5:$W$16)={v}".format(c=SH_CFG, v=code),
                     fill=bg, color=fg, bold=True))

    # ---- Diagramm ---------------------------------------------------------
    ws.row_dimensions[13].height = 8
    ws.row_dimensions[14].height = 20
    section(ws, "B14", "sec.chart", "H14")
    for r in range(15, 30):
        ws.row_dimensions[r].height = 17
    box(ws, 2, 15, 16, 29, fill=CARD)
    chart = build_chart(width=26.5, height=8.7)
    ws.add_chart(chart, "B15")

    # ---- Wochentabelle ----------------------------------------------------
    ws.row_dimensions[30].height = 10
    ws.row_dimensions[31].height = 20
    section(ws, "B31", "sec.table", "H31")
    section(ws, "J31", "sec.settings", "P31")

    ws.row_dimensions[TBL_HEAD].height = 24
    heads = [("B", "th.week"), ("C", "th.from"), ("D", "th.in"),
             ("F", "th.out"), ("G", "th.base"), ("H", "th.scn")]
    for col, key in heads:
        put(ws, "%s%d" % (col, TBL_HEAD), T(key),
            font=Font(name=FONT, size=9, bold=True, color="FFFFFF"), fill=PETROL,
            align=Alignment(horizontal="center", vertical="center"))
    put(ws, "E%d" % TBL_HEAD, None, fill=PETROL)
    ws.merge_cells("D%d:E%d" % (TBL_HEAD, TBL_HEAD))

    for i in range(12):
        r = TBL_FIRST + i
        cfg = 5 + i
        ws.row_dimensions[r].height = 18
        zebra = ZEBRA if i % 2 else CARD
        box(ws, 2, r, 8, r, fill=zebra)
        put(ws, "B%d" % r, "=%s!$Y$%d" % (SH_CFG, cfg),
            font=Font(name=FONT, size=9.5, bold=True, color=PETROL),
            align=Alignment(horizontal="center", vertical="center"))
        put(ws, "C%d" % r, "=%s!$M$%d" % (SH_CFG, cfg),
            font=Font(name=FONT, size=10, color=MUTED), fmt=FMT_DATE_DE,
            align=Alignment(horizontal="center", vertical="center"))
        ws.merge_cells("D%d:E%d" % (r, r))
        put(ws, "D%d" % r, "=%s!$O$%d" % (SH_CFG, cfg),
            font=Font(name=FONT, size=10, color=SAGE), fmt=FMT_EUR0Z_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))
        put(ws, "F%d" % r, "=-%s!$T$%d" % (SH_CFG, cfg),
            font=Font(name=FONT, size=10, color=MUTED), fmt=FMT_EUR0Z_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))
        put(ws, "G%d" % r, "=%s!$U$%d" % (SH_CFG, cfg),
            font=Font(name=FONT, size=10, color=TEXT), fmt=FMT_EUR0_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))
        put(ws, "H%d" % r, "=%s!$V$%d" % (SH_CFG, cfg),
            font=Font(name=FONT, size=10.5, bold=True, color=TEXT), fmt=FMT_EUR0_DE,
            align=Alignment(horizontal="right", vertical="center", indent=1))

    traffic_cf(ws, "B%d:B%d" % (TBL_FIRST, TBL_LAST), TBL_FIRST, chip=True)
    traffic_cf(ws, "H%d:H%d" % (TBL_FIRST, TBL_LAST), TBL_FIRST, chip=False)
    en_format(ws, "D%d:F%d" % (TBL_FIRST, TBL_LAST), FMT_EUR0Z_EN, 207)
    en_format(ws, "G%d:H%d" % (TBL_FIRST, TBL_LAST), FMT_EUR0_EN, 216)
    en_format(ws, "C%d:C%d" % (TBL_FIRST, TBL_LAST), FMT_DATE_EN, 214)

    # ---- Einstellungen ----------------------------------------------------
    box(ws, 10, TBL_HEAD, 16, 39, fill=CARD)
    settings = [
        (33, "set.openbal", OPEN_BALANCE, FMT_EUR_DE, 208, FMT_EUR_EN),
        (34, "set.mincash", MIN_RESERVE,  FMT_EUR_DE, 209, FMT_EUR_EN),
        (35, "set.yellow",  YELLOW_FACT,  "0.0",      None, None),
        (36, "set.start",   "=TODAY()-WEEKDAY(TODAY(),3)", FMT_DATE_DE, 215, FMT_DATE_EN),
        (37, "set.vat",     "=%s!$H$9" % SH_CFG,      "@",  None, None),
        (38, "set.vatday",  VAT_DAY,                  "0",  None, None),
    ]
    for row, key, value, fmt, fid, enfmt in settings:
        ws.row_dimensions[row].height = 21
        ws.merge_cells("J%d:L%d" % (row, row))
        put(ws, "J%d" % row, T(key),
            font=Font(name=FONT, size=9.5, color=TEXT),
            align=Alignment(horizontal="left", vertical="center", indent=1))
        ws.merge_cells("N%d:O%d" % (row, row))
        box(ws, 14, row, 15, row, fill=INPUT_BG, border=INPUT_BD)
        put(ws, "N%d" % row, value,
            font=Font(name=FONT, size=11, bold=True, color=PETROL),
            align=Alignment(horizontal="center", vertical="center"), fmt=fmt)
        if enfmt:
            en_format(ws, "N%d:O%d" % (row, row), enfmt, fid)

    dv_yes = DataValidation(type="list", formula1="%s!$H$8:$H$9" % SH_CFG,
                            allow_blank=False, showDropDown=False)
    dv_vday = DataValidation(type="whole", operator="between",
                             formula1="1", formula2="28", allow_blank=True)
    for dv in (dv_yes, dv_vday):
        ws.add_data_validation(dv)
    dv_yes.add("N37")
    dv_vday.add("N38")

    ws.merge_cells("J39:P39")
    put(ws, "J39", T("set.hint"),
        font=Font(name=FONT, size=8.5, italic=True, color=MUTED),
        align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[39].height = 26

    # ---- Kurzanleitung ----------------------------------------------------
    ws.row_dimensions[45].height = 10
    ws.row_dimensions[46].height = 20
    section(ws, "B46", "sec.help", "H46")
    box(ws, 2, 47, 16, 54, fill=CARD)
    for i, key in enumerate(("help.1", "help.2", "help.3", "help.4", "help.5")):
        r = 48 + i
        ws.row_dimensions[r].height = 17
        ws.merge_cells("B%d:P%d" % (r, r))
        put(ws, "B%d" % r, T(key),
            font=Font(name=FONT, size=9.5, color=TEXT),
            align=Alignment(horizontal="left", vertical="center", indent=2))
    ws.merge_cells("B53:P53")
    put(ws, "B53", T("foot.note"),
        font=Font(name=FONT, size=8.5, italic=True, color=MUTED),
        align=Alignment(horizontal="left", vertical="center", indent=2))

    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# 5f. DIAGRAMM
# ---------------------------------------------------------------------------

def axis_text(size=9, color=MUTED):
    return RichText(
        bodyPr=RichTextProperties(),
        p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=int(size * 100), b=False,
                                       solidFill=color,
                                       latin=DrawingFont(typeface=FONT)))), ])


def build_chart(width, height):
    chart = LineChart()
    chart.style = None
    chart.width = width
    chart.height = height

    for col, color, dash, w in ((21, TEAL, None, 28575),      # U  Basis
                                (22, SAGE, "dash", 25400),    # V  Szenario
                                (24, RED_FG, "sysDot", 12700)):  # X  Reserve
        ref = Reference(range_string="'{s}'!${c}$4:${c}$16".format(
            s=SH_CFG, c=get_column_letter(col)))
        chart.add_data(ref, titles_from_data=True)

    cats = Reference(range_string="'{s}'!$Y$5:$Y$16".format(s=SH_CFG))
    chart.set_categories(cats)

    styles = ((TEAL, None, 28575), (SAGE, "dash", 25400), (RED_FG, "sysDot", 12700))
    for ser, (color, dash, w) in zip(chart.series, styles):
        ser.graphicalProperties = GraphicalProperties()
        ser.graphicalProperties.line = LineProperties(solidFill=color, w=w)
        if dash:
            ser.graphicalProperties.line.prstDash = dash
        ser.smooth = False
        ser.marker = Marker(symbol="none")

    chart.graphical_properties = GraphicalProperties(solidFill="FFFFFF")
    chart.graphical_properties.line = LineProperties(noFill=True)

    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.y_axis.numFmt = FMT_EUR0_DE
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=GRIDLINE, w=6350))
    chart.x_axis.majorGridlines = None
    chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
    chart.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=HAIRLINE, w=6350))
    chart.y_axis.txPr = axis_text()
    chart.x_axis.txPr = axis_text()
    chart.y_axis.majorTickMark = "none"
    chart.x_axis.majorTickMark = "out"

    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.legend.txPr = axis_text(9, TEXT)
    return chart


# ---------------------------------------------------------------------------
# 5g. BENANNTE BEREICHE
# ---------------------------------------------------------------------------

def define_names(wb):
    names = {
        "LANG":            "%s!$E$2" % SH_CFG,
        "VATON":           "%s!$E$3" % SH_CFG,
        "OPENBAL":         "%s!$N$33" % SH_DASH,
        "MINCASH":         "%s!$N$34" % SH_DASH,
        "YELLOWF":         "%s!$N$35" % SH_DASH,
        "START":           "%s!$N$36" % SH_DASH,
        "VATDAY":          "%s!$N$38" % SH_DASH,
        "DELAY_ALL":       "%s!$J$10" % QSCN,
        "SCN_CUSTOMER":    "%s!$J$13" % QSCN,
        "DELAY_CUSTOMER":  "%s!$J$16" % QSCN,
        "LOSS_RATE":       "%s!$J$19" % QSCN,
    }
    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "Cashflow-Fruehwarnsystem.xlsx"
    print("geschrieben:", build(out))
