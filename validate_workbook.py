#!/usr/bin/env python3
"""Prueft die gebaute Arbeitsmappe: Formeln, Kennzahlen und Paket-Struktur.

Aufruf:  python3 validate_workbook.py [Datei]

Der Test rechnet alle Formeln der Mappe mit der Bibliothek ``formulas`` nach
(unabhaengige Excel-Formelmaschine) und prueft anschliessend fachliche
Invarianten gegen unabhaengig in Python berechnete Werte. Zusaetzlich wird die
OPC-Struktur der .xlsm geprueft und das eingebettete VBA-Projekt mit
``oletools`` gegengelesen.

Rueckgabewert 0 = alles in Ordnung, 1 = mindestens eine Pruefung fehlgeschlagen.
"""

from __future__ import annotations

import re
import sys
import warnings
import zipfile
from collections import defaultdict
from pathlib import Path

warnings.filterwarnings("ignore")

sys.path.insert(0, str(Path(__file__).parent / "tools"))

import beispieldaten as bd  # noqa: E402
import ovba  # noqa: E402

TOLERANZ = 1e-6

_FEHLER_MUSTER = re.compile(r"#(REF|NAME|VALUE|DIV/0|N/A|NUM|NULL)")


class Pruefung:
    def __init__(self) -> None:
        self.ok = 0
        self.fehler: list[str] = []

    def gleich(self, name: str, ist, soll, toleranz: float = TOLERANZ) -> None:
        try:
            abweichung = abs(float(ist) - float(soll))
        except (TypeError, ValueError):
            self.fehler.append(f"{name}: nicht numerisch (ist={ist!r}, soll={soll!r})")
            return
        if abweichung <= toleranz:
            self.ok += 1
            print(f"  OK   {name}: {float(ist):,.4f}")
        else:
            self.fehler.append(
                f"{name}: ist={float(ist):,.6f} soll={float(soll):,.6f} "
                f"(Abweichung {abweichung:,.6f})"
            )

    def wahr(self, name: str, bedingung: bool, detail: str = "") -> None:
        if bedingung:
            self.ok += 1
            print(f"  OK   {name}{(' – ' + detail) if detail else ''}")
        else:
            self.fehler.append(f"{name}: {detail or 'Bedingung nicht erfuellt'}")


def zellwert(loesung: dict, blatt: str, adresse: str):
    """Liest einen berechneten Wert aus dem Ergebnis von formulas."""
    # formulas schreibt den Blattnamen in Grossbuchstaben, den Dateinamen nicht
    schluessel = f"'[{loesung['__datei__']}]{blatt.upper()}'!{adresse.upper()}"
    wert = loesung.get(schluessel)
    if wert is None:
        raise KeyError(f"{blatt}!{adresse} nicht im Ergebnis (Schluessel {schluessel})")
    try:
        return wert.value[0, 0]
    except (AttributeError, TypeError, IndexError):
        return wert


def rechne(datei: Path) -> dict:
    import formulas

    print(f"Formeln werden nachgerechnet ({datei.name}) …")
    modell = formulas.ExcelModel().loads(str(datei)).finish()
    loesung = dict(modell.calculate())
    loesung["__datei__"] = datei.name
    print(f"  {len(loesung)} Zellen/Bereiche berechnet")
    return loesung


def pruefe_fehlerwerte(loesung: dict, p: Pruefung) -> None:
    """Sucht Excel-Fehlerwerte in allen berechneten Zellen."""
    treffer: dict[str, list[str]] = defaultdict(list)
    for schluessel, wert in loesung.items():
        if schluessel == "__datei__":
            continue
        try:
            roh = wert.value[0, 0]
        except (AttributeError, TypeError, IndexError):
            roh = wert
        text = str(roh)
        gefunden = _FEHLER_MUSTER.search(text)
        if gefunden:
            treffer[gefunden.group(0)].append(schluessel)

    if not treffer:
        p.wahr("keine Excel-Fehlerwerte in der Mappe", True)
        return
    for art, zellen in sorted(treffer.items()):
        p.fehler.append(
            f"{len(zellen)} Zelle(n) mit {art}, z.B. {', '.join(zellen[:6])}"
        )


def erwartungswerte() -> dict:
    """Berechnet die Sollwerte unabhaengig von Excel direkt aus den Rohdaten."""
    hist, fc = bd.historie(), bd.forecast()
    m3 = 1.5

    def palaeq(zeile: list) -> float:
        return zeile[7] if zeile[7] > 0 else (zeile[6] / m3 if m3 else 0.0)

    def tage(zeilen: list) -> int:
        return (max(z[0] for z in zeilen) - min(z[0] for z in zeilen)).days + 1

    werte = {
        "hist_zeilen": len(hist),
        "fc_zeilen": len(fc),
        "hist_tage": tage(hist),
        "fc_tage": tage(fc),
        "hist_palaeq": sum(palaeq(z) for z in hist),
        "hist_umsatz": sum(z[5] for z in hist),
        "kategorie": {},
        "region": {},
    }

    gew_hist = 0.6  # Blatt Zielreichweite, Zelle B6
    for kat, reichweite, *_ in bd.KATEGORIEN:
        h = sum(palaeq(z) for z in hist if z[3] == kat) / werte["hist_tage"]
        f = sum(palaeq(z) for z in fc if z[3] == kat) / werte["fc_tage"]
        basis = gew_hist * h + (1 - gew_hist) * f
        werte["kategorie"][kat] = {
            "hist_rate": h,
            "fc_rate": f,
            "basis_rate": basis,
            "reichweite": reichweite,
            "zielbestand": round(basis * reichweite, 0),
        }

    kat_sim = bd.KATEGORIEN[2][0]  # Vorbelegung des Blatts Simulation
    for region, _land, lat, lon in bd.REGIONEN:
        h = sum(palaeq(z) for z in hist if z[3] == kat_sim and z[2] == region) / werte["hist_tage"]
        f = sum(palaeq(z) for z in fc if z[3] == kat_sim and z[2] == region) / werte["fc_tage"]
        werte["region"][region] = {
            "basis_rate": gew_hist * h + (1 - gew_hist) * f,
            "lat": lat,
            "lon": lon,
        }
    werte["sim_kategorie"] = kat_sim
    return werte


def pruefe_fachlich(loesung: dict, p: Pruefung) -> None:
    soll = erwartungswerte()
    kat = soll["sim_kategorie"]

    print("\nZielreichweite")
    p.gleich("Bewegungszeilen Historie", zellwert(loesung, "Zielreichweite", "B10"),
             soll["hist_zeilen"])
    p.gleich("Tage im Historie-Zeitraum", zellwert(loesung, "Zielreichweite", "B8"),
             soll["hist_tage"])
    p.gleich("Tage im Forecast-Zeitraum", zellwert(loesung, "Zielreichweite", "B9"),
             soll["fc_tage"])
    for i, (name, *_rest) in enumerate(bd.KATEGORIEN):
        zeile = 14 + i
        erwartet = soll["kategorie"][name]
        p.gleich(f"Ø Bedarf/Tag Historie {name}",
                 zellwert(loesung, "Zielreichweite", f"C{zeile}"), erwartet["hist_rate"], 1e-4)
        p.gleich(f"Ø Bedarf/Tag Basis {name}",
                 zellwert(loesung, "Zielreichweite", f"E{zeile}"), erwartet["basis_rate"], 1e-4)
        p.gleich(f"Ziel-Bestand {name}",
                 zellwert(loesung, "Zielreichweite", f"F{zeile}"), erwartet["zielbestand"], 0.5)

    print("\nSimulation – Eingaben und Bedarfsschwerpunkt")
    p.gleich("Summe der Gewichtungen", zellwert(loesung, "Simulation", "B13"), 1.0)
    p.gleich("Ziel-Bestand der gewaehlten Kategorie",
             zellwert(loesung, "Simulation", "B8"),
             soll["kategorie"][kat]["zielbestand"], 0.5)

    # Der Regionsraster aggregiert unabhaengig vom Blatt Zielreichweite; beide
    # Wege muessen denselben Tagesbedarf ergeben.
    summe_regionen = zellwert(loesung, "Simulation", "I15")
    p.gleich("Bedarf/Tag: Regionsraster == Zielreichweite", summe_regionen,
             zellwert(loesung, "Simulation", "B7"), 1e-4)

    gewicht = sum(r["basis_rate"] for r in soll["region"].values())
    lat_soll = sum(r["basis_rate"] * r["lat"] for r in soll["region"].values()) / gewicht
    lon_soll = sum(r["basis_rate"] * r["lon"] for r in soll["region"].values()) / gewicht
    p.gleich("Bedarfsschwerpunkt Latitude", zellwert(loesung, "Simulation", "J15"),
             lat_soll, 1e-4)
    p.gleich("Bedarfsschwerpunkt Longitude", zellwert(loesung, "Simulation", "K15"),
             lon_soll, 1e-4)

    print("\nSimulation – Scoring je DC")
    import math

    anzahl = len(bd.DCS)
    scores, anteile, distanzen = [], [], []
    for i, dc in enumerate(bd.DCS):
        zeile = 20 + i
        name = dc[0]
        # Distanz unabhaengig nachrechnen (Orthodrome)
        d_soll = 6371 * math.acos(
            min(1.0,
                math.cos(math.radians(dc[2])) * math.cos(math.radians(lat_soll))
                * math.cos(math.radians(lon_soll) - math.radians(dc[3]))
                + math.sin(math.radians(dc[2])) * math.sin(math.radians(lat_soll)))
        )
        p.gleich(f"Distanz {name} (km)", zellwert(loesung, "Simulation", f"H{zeile}"),
                 round(d_soll), 1.0)

        auslastung = (dc[8] + soll["kategorie"][kat]["zielbestand"]) / dc[4]
        p.gleich(f"Auslastung {name}", zellwert(loesung, "Simulation", f"F{zeile}"),
                 auslastung, 1e-6)

        kosten_soll = dc[6] + dc[7] * round(d_soll)
        p.gleich(f"Transportkosten {name} (€/Pal.)",
                 zellwert(loesung, "Simulation", f"I{zeile}"), kosten_soll, 0.05)

        scores.append(float(zellwert(loesung, "Simulation", f"P{zeile}")))
        anteile.append(float(zellwert(loesung, "Simulation", f"T{zeile}")))
        distanzen.append(float(zellwert(loesung, "Simulation", f"H{zeile}")))

    print("\nSimulation – Scores, Rangfolge und Aufteilung")
    for i, dc in enumerate(bd.DCS):
        zeile = 20 + i
        for spalte, bezeichnung in (("M", "Kapazitaet"), ("N", "Transport"), ("O", "Reichweite")):
            wert = float(zellwert(loesung, "Simulation", f"{spalte}{zeile}"))
            p.wahr(f"Teil-Score {bezeichnung} {dc[0]} in [0,1]",
                   -TOLERANZ <= wert <= 1 + TOLERANZ, f"{wert:.3f}")

    gewichte = [float(zellwert(loesung, "Simulation", f"B{z}")) for z in (10, 11, 12)]
    for i, dc in enumerate(bd.DCS):
        zeile = 20 + i
        teile = [float(zellwert(loesung, "Simulation", f"{s}{zeile}")) for s in "MNO"]
        p.gleich(f"Gesamt-Score {dc[0]} = gewichtete Summe", scores[i],
                 sum(g * t for g, t in zip(gewichte, teile)), 1e-9)

    p.wahr("Teil-Score Kapazitaet: mindestens ein DC hat 1,0",
           any(abs(float(zellwert(loesung, "Simulation", f"M{20 + i}")) - 1) < 1e-9
               for i in range(anzahl)), "Min-Max-Normierung erreicht die Obergrenze")

    raenge = [zellwert(loesung, "Simulation", f"R{20 + i}") for i in range(anzahl)]
    raenge_int = sorted(int(r) for r in raenge if str(r).strip() != "")
    p.wahr("Rangfolge ist eine Permutation 1..n", raenge_int == list(range(1, anzahl + 1)),
           f"Raenge {raenge_int}")

    bester = max(range(anzahl), key=lambda i: scores[i])
    rang_des_besten = int(zellwert(loesung, "Simulation", f"R{20 + bester}"))
    p.wahr("Rang 1 gehoert zum hoechsten Gesamt-Score", rang_des_besten == 1,
           f"{bd.DCS[bester][0]} mit Score {scores[bester]:.3f}")

    max_dcs = int(zellwert(loesung, "Simulation", "B16"))
    p.gleich("Summe der empfohlenen Anteile (Splitting)", sum(anteile), 1.0, 1e-9)
    p.wahr("nur die besten n DCs erhalten einen Anteil",
           sum(1 for a in anteile if a > 0) == max_dcs,
           f"{sum(1 for a in anteile if a > 0)} von {anzahl} DCs bei max. {max_dcs}")

    print("\nErgebniszeile und Dashboard")
    text = str(zellwert(loesung, "Simulation", "A28"))
    p.wahr("Ergebniszeile nennt Empfehlung und Anteile",
           "Empfehlung" in text and "%" in text, text[:110])
    p.wahr("Ergebniszeile nennt den bestplatzierten DC",
           bd.DCS[bester][0] in text, bd.DCS[bester][0])

    p.gleich("KPI Paletten-Äquivalent Historie",
             zellwert(loesung, "Dashboard", "F8"), soll["hist_palaeq"], 0.01)
    p.gleich("KPI Umsatz Historie", zellwert(loesung, "Dashboard", "J8"),
             soll["hist_umsatz"], 0.5)
    p.gleich("KPI Bewegungszeilen", zellwert(loesung, "Dashboard", "B8"),
             soll["hist_zeilen"])

    monate = [float(zellwert(loesung, "Dashboard", f"C{z}")) for z in range(51, 63)]
    p.gleich("Monatsraster summiert die Historie vollstaendig", sum(monate),
             soll["hist_palaeq"], 0.01)
    p.wahr("alle zwoelf Monatswerte sind belegt", all(m > 0 for m in monate),
           f"min={min(monate):,.1f} max={max(monate):,.1f}")

    p.wahr("Distanzen sind plausibel (50–2500 km)",
           all(50 <= d <= 2500 for d in distanzen),
           f"{min(distanzen):,.0f}–{max(distanzen):,.0f} km")

    band = str(zellwert(loesung, "Dashboard", "B15"))
    p.wahr("Ergebnisband des Dashboards nennt Kategorie und Empfehlung",
           kat in band and bd.DCS[bester][0] in band, band[:110])

    print("\nKarte")
    for i, dc in enumerate(bd.DCS):
        zeile = 7 + i
        p.gleich(f"Karte Longitude {dc[0]}", zellwert(loesung, "Karte", f"B{zeile}"),
                 dc[3], 1e-9)
        p.gleich(f"Karte Kapazitaet {dc[0]}", zellwert(loesung, "Karte", f"D{zeile}"),
                 dc[4], 1e-9)
    for i, (region, _land, _lat, lon) in enumerate(bd.REGIONEN):
        zeile = 7 + i
        p.gleich(f"Karte Bedarf {region}", zellwert(loesung, "Karte", f"I{zeile}"),
                 soll["region"][region]["basis_rate"], 1e-4)
        p.gleich(f"Karte Longitude {region}", zellwert(loesung, "Karte", f"G{zeile}"),
                 lon, 1e-9)

    print("\nReichweiten-Kriterium")
    reichweite_roh = [float(zellwert(loesung, "Simulation", f"L{20 + i}"))
                      for i in range(anzahl)]
    p.wahr("Reichweiten-Score unterscheidet die DCs",
           max(reichweite_roh) - min(reichweite_roh) > 1e-6,
           f"{min(reichweite_roh):.3f} bis {max(reichweite_roh):.3f}")
    eng = [i for i in range(anzahl)
           if bd.DCS[i][4] - bd.DCS[i][8] < soll["kategorie"][kat]["zielbestand"]]
    p.wahr("kapazitaetsbeschraenkte DCs erhalten einen Abschlag",
           all(reichweite_roh[i] < 1 - 1e-9 for i in eng),
           ", ".join(bd.DCS[i][0] for i in eng) or "keiner")


def pruefe_paket(datei: Path, p: Pruefung) -> None:
    print("\nPaket-Struktur und VBA-Projekt")
    with zipfile.ZipFile(datei) as z:
        namen = z.namelist()
        ct = z.read("[Content_Types].xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        vba = z.read("xl/vbaProject.bin")

    p.wahr("xl/vbaProject.bin ist enthalten", "xl/vbaProject.bin" in namen,
           f"{len(vba):,} Byte")
    p.wahr("Workbook-Part ist als macroEnabled deklariert",
           "application/vnd.ms-excel.sheet.macroEnabled.main+xml" in ct)
    p.wahr("Content-Type fuer .bin vorhanden",
           'Extension="bin"' in ct and "vbaProject" in ct)
    p.wahr("Beziehung auf vbaProject.bin vorhanden", "vbaProject.bin" in rels)
    p.wahr("fullCalcOnLoad gesetzt (Excel rechnet beim Oeffnen)",
           'fullCalcOnLoad="1"' in
           zipfile.ZipFile(datei).read("xl/workbook.xml").decode("utf-8"))

    # VBA-Projekt mit oletools gegenlesen
    import io

    import olefile
    from oletools.olevba import decompress_stream

    ole = olefile.OleFileIO(io.BytesIO(vba))
    module = sorted(e[1] for e in ole.listdir() if e[0] == "VBA"
                    and e[1] not in ("_VBA_PROJECT", "dir"))
    p.wahr("VBA-Module im Container", len(module) == 8, ", ".join(module))

    projekt = ole.openstream("PROJECT").read().decode("latin-1")
    for feld in ("CMG", "DPB", "GC"):
        wert = re.search(rf'{feld}="([0-9A-F]+)"', projekt)
        p.wahr(f"PROJECT-Feld {feld} entschluesselbar", wert is not None)
        if wert:
            version, key, klartext = ovba.decrypt_hex(wert.group(1))
            p.wahr(f"  {feld}: Version 2, Klartext {klartext.hex()}",
                   version == 2 and key == ovba.project_key(
                       re.search(r'ID="(\{[^}]+\})"', projekt).group(1)))

    gesamt_quelltext = 0
    for name in module:
        quelle = ovba.decompress(ole.openstream(["VBA", name]).read()).decode("latin-1")
        p.wahr(f"Modul {name} lesbar", quelle.startswith("Attribute VB_Name"),
               f"{len(quelle):,} Zeichen")
        gesamt_quelltext += len(quelle)
        # oletools als zweite, unabhaengige Dekompression
        vergleich = bytes(decompress_stream(
            bytearray(ole.openstream(["VBA", name]).read()))).decode("latin-1")
        p.wahr(f"  {name}: oletools liest identischen Quelltext", vergleich == quelle)
    print(f"  VBA-Quelltext insgesamt: {gesamt_quelltext:,} Zeichen")
    ole.close()


def pruefe_struktur(datei: Path, p: Pruefung) -> None:
    """Prueft Bestandteile, die keine Formelmaschine sehen kann."""
    from openpyxl import load_workbook

    print("\nBlatt-Struktur, Tabellen, Prüfregeln und Diagramme")
    wb = load_workbook(datei, keep_vba=True)

    erwartete_blaetter = [
        "Dashboard", "DC_Stammdaten", "Import_Historie", "Import_Forecast",
        "Import_Mapping", "Zielreichweite", "Simulation", "Szenarien", "Karte",
        "Beispieldaten",
    ]
    p.wahr("alle zehn Blaetter in der vorgesehenen Reihenfolge",
           wb.sheetnames == erwartete_blaetter, ", ".join(wb.sheetnames))

    p.wahr("Gitternetzlinien auf allen Blaettern ausgeblendet",
           all(not ws.sheet_view.showGridLines for ws in wb.worksheets),
           ", ".join(ws.title for ws in wb.worksheets
                     if ws.sheet_view.showGridLines) or "keine Ausnahme")

    tabellen = {name: ws.title for ws in wb.worksheets for name in ws.tables}
    for name, blatt in [("tblDC", "DC_Stammdaten"), ("tblRegionen", "DC_Stammdaten"),
                        ("tblHistorie", "Import_Historie"),
                        ("tblForecast", "Import_Forecast"),
                        ("tblMapping", "Import_Mapping"),
                        ("tblZielreichweite", "Zielreichweite")]:
        p.wahr(f"Tabellenobjekt {name} auf {blatt}", tabellen.get(name) == blatt,
               wb[blatt].tables[name].ref if name in wb[blatt].tables else "fehlt")

    # Die Spaltenkoepfe der Mapping-Tabelle liest der M-Code namentlich.
    mapping_kopf = [wb["Import_Mapping"].cell(11, c).value for c in range(1, 5)]
    p.wahr("Mapping-Spaltenkoepfe passen zum M-Code",
           mapping_kopf == ["Zielfeld", "Quellspalte Historie",
                            "Quellspalte Forecast", "Hinweis"],
           str(mapping_kopf))

    sim = wb["Simulation"]
    dv_bereiche = {str(dv.sqref): dv.type for dv in sim.data_validations.dataValidation}
    p.wahr("Dropdown fuer die Produktkategorie", "B4" in dv_bereiche)
    p.wahr("Gewichtungen 0–100 % per Datenprüfung begrenzt",
           dv_bereiche.get("B10:B12") == "decimal")
    p.wahr("Dropdown fuer den Zuordnungsmodus", dv_bereiche.get("B15") == "list")
    p.wahr("bedingte Formatierung auf den Score-Spalten",
           len([r for r in sim.conditional_formatting]) >= 4,
           f"{len([r for r in sim.conditional_formatting])} Bereiche")
    p.wahr("Ergebniszeile ist verbunden und hervorgehoben",
           any(str(b).startswith("A28:") for b in sim.merged_cells.ranges),
           ", ".join(str(b) for b in sim.merged_cells.ranges
                     if str(b).startswith("A28")))

    dc = wb["DC_Stammdaten"]
    p.wahr("Ja/Nein-Prüfregel auf der Spalte aktiv",
           any(dv.type == "list" and "J" in str(dv.sqref)
               for dv in dc.data_validations.dataValidation))

    diagramme = {ws.title: len(ws._charts) for ws in wb.worksheets}
    p.wahr("vier Diagramme auf dem Dashboard", diagramme.get("Dashboard") == 4,
           str(diagramme))
    p.wahr("Blasendiagramm auf dem Blatt Karte", diagramme.get("Karte") == 1)
    p.wahr("Vergleichsdiagramm auf dem Blatt Szenarien",
           diagramme.get("Szenarien") == 1)

    namen = set(wb.defined_names)
    p.wahr("benannte Bereiche fuer Power Query vorhanden",
           {"Pfad_Historie", "Pfad_Forecast", "Param_Trennzeichen",
            "Param_M3_pro_Palette", "Param_Quelltabelle"} <= namen,
           ", ".join(sorted(namen)))

    stile = set(wb.named_styles)
    p.wahr("einheitliche Zellstile registriert",
           len([s for s in stile if s.startswith("LNP_")]) >= 20,
           f"{len([s for s in stile if s.startswith('LNP_')])} Stile")


def main() -> int:
    datei = Path(sys.argv[1] if len(sys.argv) > 1 else "Logistik_Netzwerkplanung.xlsm")
    if not datei.exists():
        print(f"Datei nicht gefunden: {datei}")
        return 1

    p = Pruefung()
    pruefe_paket(datei, p)
    pruefe_struktur(datei, p)
    loesung = rechne(datei)
    pruefe_fehlerwerte(loesung, p)
    pruefe_fachlich(loesung, p)

    print("\n" + "=" * 72)
    if p.fehler:
        print(f"FEHLGESCHLAGEN: {len(p.fehler)} Problem(e), {p.ok} Pruefungen ok")
        for eintrag in p.fehler:
            print(f"  FEHLER  {eintrag}")
        return 1
    print(f"ALLE {p.ok} PRUEFUNGEN ERFOLGREICH")
    return 0


if __name__ == "__main__":
    sys.exit(main())
