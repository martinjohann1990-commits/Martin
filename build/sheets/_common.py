"""Small layout helpers shared by every sheet module.

Kept deliberately dumb: these only place and style cells. Calculation
content is added sheet by sheet starting with Milestone 2/3.
"""

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from build import styles


def init_sheet(ws: Worksheet, tab_color=None, show_gridlines=False, freeze="A3"):
    ws.sheet_view.showGridLines = show_gridlines
    styles.apply_tab_color(ws, tab_color)
    if freeze:
        ws.freeze_panes = freeze


def write_heading(ws: Worksheet, text: str, row: int = 2, col: int = 2):
    cell = ws.cell(row=row, column=col, value=text)
    styles.style_title(cell)
    return row + 1


def write_intro(ws: Worksheet, text: str, row: int, col: int = 2, span: int = 8, height: int = 40):
    cell = ws.cell(row=row, column=col, value=text)
    styles.style_intro(cell)
    end_col = get_column_letter(col + span - 1)
    ws.merge_cells(f"{get_column_letter(col)}{row}:{end_col}{row}")
    ws.row_dimensions[row].height = height
    return row + 1


def write_note(ws: Worksheet, text: str, row: int, col: int = 2, span: int = 8):
    cell = ws.cell(row=row, column=col, value=text)
    styles.style_note(cell)
    end_col = get_column_letter(col + span - 1)
    ws.merge_cells(f"{get_column_letter(col)}{row}:{end_col}{row}")
    return row + 1


def write_section_header(ws: Worksheet, text: str, row: int, col: int = 2, span: int = 8):
    cell = ws.cell(row=row, column=col, value=text)
    styles.style_section_header(cell)
    end_col = get_column_letter(col + span - 1)
    ws.merge_cells(f"{get_column_letter(col)}{row}:{end_col}{row}")
    ws.row_dimensions[row].height = 20
    return row + 1


def write_table_headers(ws: Worksheet, row: int, start_col: int, headers: list):
    for i, text in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=text)
        styles.style_table_header(cell)
    ws.row_dimensions[row].height = 28
    return row + 1


def write_label_value_row(ws: Worksheet, row: int, label_col: int, value_col: int,
                           label: str, unit: str = None, unit_col: int = None,
                           number_format: str = None):
    """One Settings-style input row: label cell + empty, unlocked value cell."""
    lc = ws.cell(row=row, column=label_col, value=label)
    styles.style_label(lc)
    vc = ws.cell(row=row, column=value_col)
    styles.style_input(vc, number_format=number_format)
    if unit:
        uc = ws.cell(row=row, column=unit_col or (value_col + 1), value=unit)
        styles.style_note(uc)
    return row + 1
