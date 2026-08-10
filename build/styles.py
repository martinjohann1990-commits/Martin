"""Shared colors, fonts, number formats and cell styles.

Every sheet module pulls its look from here so the workbook reads as one
product instead of twelve separately-styled sheets. The only style
distinction that matters functionally is input vs. locked: input cells
are unlocked and visually distinct so a buyer never has to guess where to
type.
"""

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

COLORS = {
    "primary": "1F4E4C",
    "primary_dark": "163936",
    "accent": "C9A227",
    "background": "FFFFFF",
    "background_alt": "F7F5F0",
    "input_fill": "FFF6DA",
    "input_border": "C9A227",
    "locked_text": "222222",
    "muted_text": "6B6B6B",
    "header_text": "FFFFFF",
    "warning": "B00020",
    "warning_fill": "FBE1E1",
    "ok": "2E7D32",
    "ok_fill": "E4F2E4",
    "grid": "D9D5C9",
}

NUMBER_FORMATS = {
    "currency": '#,##0.00',
    "currency_rounded": '#,##0',
    "percent": '0%',
    "percent_1dp": '0.0%',
    "integer": '#,##0',
    "decimal": '#,##0.00',
    "date": 'yyyy-mm-dd',
    "week": '"W"0',
}

FONT_NAME = "Calibri"


def font_title():
    return Font(name=FONT_NAME, size=16, bold=True, color=COLORS["primary"])


def font_section_header():
    return Font(name=FONT_NAME, size=12, bold=True, color=COLORS["header_text"])


def font_table_header():
    return Font(name=FONT_NAME, size=10, bold=True, color=COLORS["header_text"])


def font_label():
    return Font(name=FONT_NAME, size=10, bold=False, color=COLORS["locked_text"])


def font_label_bold():
    return Font(name=FONT_NAME, size=10, bold=True, color=COLORS["locked_text"])


def font_input():
    return Font(name=FONT_NAME, size=10, bold=False, color=COLORS["locked_text"])


def font_formula():
    return Font(name=FONT_NAME, size=10, bold=False, color=COLORS["locked_text"])


def font_note():
    return Font(name=FONT_NAME, size=9, italic=True, color=COLORS["muted_text"])


def font_body():
    return Font(name=FONT_NAME, size=10, color=COLORS["locked_text"])


FILL_SECTION_HEADER = PatternFill("solid", fgColor=COLORS["primary"])
FILL_TABLE_HEADER = PatternFill("solid", fgColor=COLORS["primary_dark"])
FILL_INPUT = PatternFill("solid", fgColor=COLORS["input_fill"])
FILL_ALT_ROW = PatternFill("solid", fgColor=COLORS["background_alt"])
FILL_WARNING = PatternFill("solid", fgColor=COLORS["warning_fill"])
FILL_OK = PatternFill("solid", fgColor=COLORS["ok_fill"])

THIN_GRID = Border(
    left=Side(style="thin", color=COLORS["grid"]),
    right=Side(style="thin", color=COLORS["grid"]),
    top=Side(style="thin", color=COLORS["grid"]),
    bottom=Side(style="thin", color=COLORS["grid"]),
)


def style_title(cell):
    cell.font = font_title()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section_header(cell):
    cell.font = font_section_header()
    cell.fill = FILL_SECTION_HEADER
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def style_table_header(cell):
    cell.font = font_table_header()
    cell.fill = FILL_TABLE_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_GRID


def style_label(cell, bold=False):
    cell.font = font_label_bold() if bold else font_label()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_input(cell, number_format=None):
    """Mark a cell as user-editable: unlocked and visually distinct."""
    cell.font = font_input()
    cell.fill = FILL_INPUT
    cell.border = THIN_GRID
    cell.protection = Protection(locked=False)
    if number_format:
        cell.number_format = number_format


def style_formula(cell, number_format=None):
    """Mark a cell as computed: locked, plain background."""
    cell.font = font_formula()
    cell.border = THIN_GRID
    cell.protection = Protection(locked=True)
    if number_format:
        cell.number_format = number_format


def style_note(cell):
    cell.font = font_note()
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_intro(cell):
    cell.font = font_body()
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_column_widths(ws: Worksheet, widths: dict):
    """widths: {1: 12, 2: 30, ...} or {'A': 12, 'B': 30, ...}"""
    for key, width in widths.items():
        letter = key if isinstance(key, str) else get_column_letter(key)
        ws.column_dimensions[letter].width = width


def protect_sheet(ws: Worksheet):
    """Lock formula cells but allow unlocking without a password, per spec
    rule 3: advanced users must be able to remove protection themselves."""
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def apply_tab_color(ws: Worksheet, hex_color=None):
    ws.sheet_properties.tabColor = hex_color or COLORS["primary"]
