#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Design-Vorschau des erzeugten Cashflow-Tools.

Rendert die sichtbaren Blaetter als HTML nach (Fuellungen, Rahmen, Schriften,
verbundene Zellen, Spaltenbreiten, Zeilenhoehen) und loest die Formeln mit den
Werten aus der Modell-Nachrechnung auf. Dient ausschliesslich der optischen
Kontrolle - die Excel-Datei selbst bleibt unberuehrt.

Aufruf:  python3 preview_tool.py [datei.xlsx] [ausgabe.html]
"""

import html
import re
import sys
from datetime import date, timedelta

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

import build_cashflow_tool as B
from verify_tool import simulate

LANG = 1  # 1 = Deutsch, 2 = English (Vorschausprache)

TODAY = date.today()
START = TODAY - timedelta(days=TODAY.weekday())
ROWS = simulate(B.SCN_DELAY_ALL, B.SCN_DELAY_CUST, B.SCN_LOSS, B.SCN_CUSTOMER)
BASE = simulate(0, 0, 0.0, None)
TR = {k: (de, en) for k, de, en in B.TRANSLATIONS}
MAXCODE = max(r[6] for r in ROWS)


def t(key):
    return TR[key][LANG - 1]


# --- Zahlen- und Datumsformatierung nach deutschem Muster -------------------

def fmt_number(v, decimals):
    s = "%,.*f" % (decimals, v) if False else format(abs(v), ",.%df" % decimals)
    s = s.replace(",", " ").replace(".", ",").replace(" ", ".")
    return ("-" if v < 0 else "") + s


def apply_format(value, code):
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return value
    if code in (B.FMT_EUR_DE,):
        return fmt_number(value, 2) + " €"
    if code in (B.FMT_EUR0_DE,):
        return fmt_number(value, 0) + " €"
    if code in (B.FMT_EUR0Z_DE,):
        return "–" if round(value) == 0 else fmt_number(value, 0) + " €"
    if code == B.FMT_PCT:
        return fmt_number(value * 100, 0) + " %"
    if code == B.FMT_DATE_DE:
        d = value if isinstance(value, date) else None
        return d.strftime("%d.%m.%Y") if d else str(value)
    if code == "0":
        return "%d" % value
    if code == "0.0":
        return fmt_number(value, 1)
    if isinstance(value, float):
        return fmt_number(value, 2)
    return str(value)


# --- Formelaufloesung -------------------------------------------------------

CFG = {}  # (spalte, zeile) -> Wert im Rechenkern


def fill_engine():
    for i, (ws_, inb, ins, out, bb, bs, code) in enumerate(ROWS):
        r = 5 + i
        CFG[("M", r)] = ws_
        CFG[("N", r)] = ws_ + timedelta(days=6)
        CFG[("O", r)] = BASE[i][1]
        CFG[("P", r)] = ins
        CFG[("T", r)] = out
        CFG[("U", r)] = bb
        CFG[("V", r)] = bs
        CFG[("W", r)] = code
        CFG[("X", r)] = B.MIN_RESERVE
        CFG[("Y", r)] = "%s %d" % (t("lbl.week"), ws_.isocalendar()[1])


fill_engine()

OPEN_SUM = sum(a for _c, _n, a, _d, _tm, paid in B.INVOICES if not paid)
OPEN_CNT = sum(1 for *_x, paid in B.INVOICES if not paid)
NAMED = {
    "OPENBAL": B.OPEN_BALANCE, "MINCASH": B.MIN_RESERVE, "YELLOWF": B.YELLOW_FACT,
    "START": START, "VATDAY": B.VAT_DAY, "DELAY_ALL": B.SCN_DELAY_ALL,
    "DELAY_CUSTOMER": B.SCN_DELAY_CUST, "LOSS_RATE": B.SCN_LOSS,
    "SCN_CUSTOMER": B.SCN_CUSTOMER, "LANG": LANG, "VATON": 0,
}

RE_T = re.compile(r'INDEX\(Config!\$B\$\d+:\$C\$\d+,MATCH\("([a-z0-9._]+)"')
RE_CFG = re.compile(r"Config!\$([A-Z]+)\$(\d+)")
RE_TODAY = re.compile(r"^=TODAY\(\)([+-]\d+)$")


def resolve(sheet, coord, formula):
    """Loest die im Tool verwendeten Formelmuster fuer die Vorschau auf."""
    f = formula

    # Statuskachel und Warnbanner
    if "MAX(Config!$W$5:$W$16)=2" in f:
        keys = RE_T.findall(f)
        if f.startswith("=IF(MAX"):
            if "alert.green" in keys:
                if MAXCODE == 0:
                    return t("alert.green")
                warn = t("alert.red") if MAXCODE == 2 else t("alert.amber")
                first = next(i for i, r in enumerate(ROWS) if r[6] == MAXCODE)
                return "%s %s." % (warn, CFG[("Y", 5 + first)])
            order = {0: "status.green", 1: "status.amber", 2: "status.red"}
            return t(order[MAXCODE])

    # Erste Warnwoche im Szenario
    if "MATCH(MAX(Config!$W$5:$W$16)" in f and "scn.none" in f:
        if MAXCODE == 0:
            return t("scn.none")
        first = next(i for i, r in enumerate(ROWS) if r[6] == MAXCODE)
        return CFG[("Y", 5 + first)]

    # Anzahl offener Rechnungen + Text
    if f.startswith("=COUNTIF(") and "kpi.recv.sub" in f:
        return "%d %s" % (OPEN_CNT, t("kpi.recv.sub"))

    # Reine Uebersetzung
    m = RE_T.search(f)
    if m and f.startswith("=INDEX(Config!"):
        return t(m.group(1))

    # Statuswerte und Dropdownanzeige
    if f == "=Config!$H$5":
        return t("val.open")
    if f == "=Config!$H$6":
        return t("val.paid")
    if f == "=Config!$H$9":
        return t("val.no")

    # Benannte Bereiche
    if f[1:] in NAMED:
        return NAMED[f[1:]]

    # Rechenkern
    m = RE_CFG.search(f)
    if m and re.fullmatch(r"=-?Config!\$[MNOPQRSTUVWXY]\$\d+", f):
        val = CFG.get((m.group(1), int(m.group(2))))
        return -val if f.startswith("=-") and isinstance(val, (int, float)) else val

    # Aggregate
    if f == "=MIN(Config!$V$5:$V$16)":
        return min(r[5] for r in ROWS)
    if f == "=MIN(Config!$U$5:$U$16)":
        return min(r[4] for r in ROWS)
    if "MIN(Config!$V$5:$V$16)-MIN(Config!$U$5:$U$16)" in f:
        return min(r[5] for r in ROWS) - min(r[4] for r in ROWS)
    if f.startswith("=SUM(") and "$Q$9:$Q$108" in f:
        return OPEN_SUM
    if f.startswith("=SUM(") and "$S$9:$S$108" in f:
        return sum(a * (1 - (B.SCN_LOSS if c == B.SCN_CUSTOMER else 0))
                   for c, _n, a, _d, _tm, paid in B.INVOICES if not paid)
    if re.fullmatch(r"=Config!\$V\$(\d+)-Config!\$U\$(\d+)", f):
        i = int(re.findall(r"\d+", f)[0]) - 5
        return ROWS[i][5] - BASE[i][4]
    if f.startswith("=SUM(D"):
        a, b = map(int, re.findall(r"D(\d+):D(\d+)", f)[0])
        if a == B.FIX_FIRST:
            return sum(c[2] for c in B.FIXED_COSTS)
        return sum(c[2] for c in B.VAR_COSTS)

    # Datumsformeln
    if f == "=TODAY()-WEEKDAY(TODAY(),3)":
        return START
    m = RE_TODAY.match(f)
    if m:
        return TODAY + timedelta(days=int(m.group(1)))

    # Rechnungszeilen
    mm = re.search(r"\$([A-Z])(\d+)", f)
    if sheet == B.SH_INV and mm:
        row = int(mm.group(2))
        idx = row - 9
        if 0 <= idx < len(B.INVOICES):
            cust, no, amount, d_inv, terms, paid = B.INVOICES[idx]
            col = coord[0]
            inv_date = TODAY + timedelta(days=d_inv)
            due = inv_date + timedelta(days=terms)
            if col == "G":
                return amount
            if col == "J":
                return due
            if col == "L":
                if paid:
                    return ""
                extra = B.SCN_DELAY_CUST if cust == B.SCN_CUSTOMER else 0
                return max(due, START) + timedelta(days=B.SCN_DELAY_ALL + extra)
            if col == "M":
                if paid:
                    return ""
                return amount * (1 - (B.SCN_LOSS if cust == B.SCN_CUSTOMER else 0))
        return ""

    return ""


# --- HTML-Erzeugung ---------------------------------------------------------

def rgb(color):
    if color is None:
        return None
    v = getattr(color, "rgb", None)
    if not isinstance(v, str) or len(v) < 6:
        return None
    return "#" + v[-6:]


def cf_style(sheet, col, row):
    """Bildet die bedingte Formatierung nach, so wie der Anwender sie sieht."""
    codes = {0: ("#" + B.GREEN_BG, "#" + B.GREEN_FG),
             1: ("#" + B.AMBER_BG, "#" + B.AMBER_FG),
             2: ("#" + B.RED_BG, "#" + B.RED_FG)}
    if sheet == B.SH_DASH:
        if row in range(B.TBL_FIRST, B.TBL_LAST + 1) and col in ("B", "H"):
            bg, fg = codes[ROWS[row - B.TBL_FIRST][6]]
            return (bg if col == "B" else None), fg, True
        if row == 6 and col in ("N", "O", "P"):
            return codes[MAXCODE][1], None, False
        if row == 8 and col in ("N", "O", "P"):
            return None, codes[MAXCODE][1], True
        if row == 12:
            bg, fg = codes[MAXCODE]
            return bg, fg, True
        if row == 8 and col in ("J", "K", "L") and min(r[5] for r in ROWS) < B.MIN_RESERVE:
            return None, "#" + B.RED_FG, True
    if sheet == B.SH_SCN:
        if 32 <= row <= 43 and col in ("B", "D"):
            bg, fg = codes[ROWS[row - 32][6]]
            return (bg if col == "B" else None), fg, True
    if sheet == B.SH_INV and 9 <= row <= 108:
        idx = row - 9
        if idx < len(B.INVOICES):
            cust, no, amount, d_inv, terms, paid = B.INVOICES[idx]
            if not paid and TODAY + timedelta(days=d_inv + terms) < TODAY:
                return "#" + B.RED_BG, "#" + B.RED_FG, False
    return None, None, False


def render_sheet(ws, max_col, max_row, charts):
    merged = {}
    skip = set()
    for m in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(m))
        merged[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    skip.add((r, c))

    widths, xoff = [], [0]
    for c in range(1, max_col + 1):
        d = ws.column_dimensions.get(get_column_letter(c))
        w = d.width if d and d.width else 8.43
        px = round(w * 7) + 5
        widths.append(px)
        xoff.append(xoff[-1] + px)

    heights, yoff = [], [0]
    for r in range(1, max_row + 1):
        d = ws.row_dimensions.get(r)
        h = d.height if d and d.height else 15
        px = round(h * 96 / 72)
        heights.append(px)
        yoff.append(yoff[-1] + px)

    out = ['<div class="sheetwrap" style="position:relative;width:%dpx">' % xoff[-1]]
    out.append('<table class="sheet"><colgroup>')
    out += ['<col style="width:%dpx">' % w for w in widths]
    out.append("</colgroup><tbody>")

    for r in range(1, max_row + 1):
        out.append('<tr style="height:%dpx">' % heights[r - 1])
        for c in range(1, max_col + 1):
            if (r, c) in skip:
                continue
            cell = ws.cell(row=r, column=c)
            col = get_column_letter(c)
            span = merged.get((r, c))
            attrs = ""
            if span:
                if span[0] > 1:
                    attrs += ' rowspan="%d"' % span[0]
                if span[1] > 1:
                    attrs += ' colspan="%d"' % span[1]

            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                v = resolve(ws.title, "%s%d" % (col, r), v)
            text = apply_format(v, cell.number_format)

            st = []
            bg = rgb(cell.fill.fgColor) if cell.fill and cell.fill.fill_type else None
            fg = rgb(cell.font.color) if cell.font else None
            bold = bool(cell.font.b) if cell.font else False
            cbg, cfg_, cbold = cf_style(ws.title, col, r)
            if cbg:
                bg = cbg
            if cfg_:
                fg = cfg_
            if cbold:
                bold = True
            if bg and bg != "#000000":
                st.append("background:%s" % bg)
            if fg:
                st.append("color:%s" % fg)
            if cell.font and cell.font.sz:
                st.append("font-size:%.1fpx" % (float(cell.font.sz) * 96 / 72))
            if bold:
                st.append("font-weight:600")
            if cell.font and cell.font.i:
                st.append("font-style:italic")
            al = cell.alignment
            if al:
                st.append("text-align:%s" % (al.horizontal or "left"))
                st.append("vertical-align:%s" % {"center": "middle", "top": "top",
                                                 "bottom": "bottom"}.get(al.vertical, "middle"))
                if al.indent:
                    st.append("padding-left:%dpx" % (al.indent * 5))
                if al.wrap_text:
                    st.append("white-space:normal")
            bd = cell.border
            for side, css in (("left", "border-left"), ("right", "border-right"),
                              ("top", "border-top"), ("bottom", "border-bottom")):
                s = getattr(bd, side, None)
                if s is not None and s.style:
                    st.append("%s:1px solid %s" % (css, rgb(s.color) or "#DDE6EC"))
            out.append("<td%s style=\"%s\">%s</td>" % (attrs, ";".join(st),
                                                       html.escape(str(text))))
        out.append("</tr>")
    out.append("</tbody></table>")

    for anchor, w_cm, h_cm in charts:
        c0 = openpyxl.utils.cell.coordinate_to_tuple(anchor)
        top = yoff[c0[0] - 1]
        left = xoff[c0[1] - 1]
        out.append('<div style="position:absolute;left:%dpx;top:%dpx;'
                   'width:%dpx;height:%dpx">%s</div>'
                   % (left, top, round(w_cm * 96 / 2.54), round(h_cm * 96 / 2.54),
                      svg_chart(round(w_cm * 96 / 2.54), round(h_cm * 96 / 2.54))))
    out.append("</div>")
    return "\n".join(out)


def svg_chart(w, h):
    """Naeherung des Liniendiagramms zur optischen Kontrolle."""
    pad_l, pad_r, pad_t, pad_b = 62, 12, 14, 46
    iw, ih = w - pad_l - pad_r, h - pad_t - pad_b
    base = [r[4] for r in BASE]
    scn = [r[5] for r in ROWS]
    lo = min(min(base), min(scn), B.MIN_RESERVE)
    hi = max(max(base), max(scn), B.MIN_RESERVE)
    lo, hi = min(0, lo), hi * 1.08

    def X(i):
        return pad_l + iw * i / 11.0

    def Y(v):
        return pad_t + ih - ih * (v - lo) / (hi - lo)

    def path(series):
        return " ".join(("M" if i == 0 else "L") + "%.1f %.1f" % (X(i), Y(v))
                        for i, v in enumerate(series))

    parts = ['<svg width="%d" height="%d" style="background:#fff;font-family:Segoe UI,Arial">' % (w, h)]
    steps = 5
    for k in range(steps + 1):
        v = lo + (hi - lo) * k / steps
        y = Y(v)
        parts.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" stroke="%s" stroke-width="1"/>'
                     % (pad_l, y, w - pad_r, y, "#" + B.GRIDLINE))
        parts.append('<text x="%d" y="%.1f" font-size="10" fill="#%s" text-anchor="end">%s €</text>'
                     % (pad_l - 7, y + 3, B.MUTED, fmt_number(v, 0)))
    parts.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" stroke="#%s"/>'
                 % (pad_l, Y(lo), w - pad_r, Y(lo), B.HAIRLINE))
    parts.append('<path d="%s" fill="none" stroke="#%s" stroke-width="2.3"/>' % (path(base), B.TEAL))
    parts.append('<path d="%s" fill="none" stroke="#%s" stroke-width="2" stroke-dasharray="7 4"/>'
                 % (path(scn), B.SAGE))
    parts.append('<path d="%s" fill="none" stroke="#%s" stroke-width="1.2" stroke-dasharray="2 3"/>'
                 % (path([B.MIN_RESERVE] * 12), B.RED_FG))
    for i in range(12):
        parts.append('<text x="%.1f" y="%d" font-size="9.5" fill="#%s" text-anchor="middle">%s</text>'
                     % (X(i), pad_t + ih + 16, B.MUTED, CFG[("Y", 5 + i)]))
    lx = pad_l + 10
    for label, color, dash in ((t("ser.base"), B.TEAL, ""),
                               (t("ser.scn"), B.SAGE, '7 4'),
                               (t("ser.min"), B.RED_FG, '2 3')):
        parts.append('<line x1="%d" y1="%d" x2="%d" y2="%d" stroke="#%s" stroke-width="2" %s/>'
                     % (lx, h - 12, lx + 22, h - 12, color,
                        'stroke-dasharray="%s"' % dash if dash else ""))
        parts.append('<text x="%d" y="%d" font-size="10" fill="#%s">%s</text>'
                     % (lx + 28, h - 8, B.TEXT, html.escape(label)))
        lx += 34 + len(label) * 6.4
    parts.append("</svg>")
    return "".join(parts)


CSS = """
body{margin:0;padding:24px;background:#dfe6ea;font-family:'Segoe UI',Arial,sans-serif}
h2{font:600 13px 'Segoe UI';color:#456;margin:26px 0 8px;letter-spacing:.08em;text-transform:uppercase}
table.sheet{border-collapse:collapse;table-layout:fixed}
table.sheet td{overflow:hidden;white-space:nowrap;font-size:13px;color:#1E2A32;
  padding:0 3px;line-height:1.15}
.sheetwrap{box-shadow:0 6px 24px rgba(20,50,70,.16);background:#fff}
"""


def main(xlsx, out_html):
    wb = openpyxl.load_workbook(xlsx)
    plan = [
        (B.SH_DASH, 17, 54, [("B15", 25.4, 8.7)]),
        (B.SH_SCN, 17, 43, [("H31", 15.0, 7.0)]),
        (B.SH_INV, 14, 28, []),
        (B.SH_COST, 6, 66, []),
    ]
    body = ["<style>%s</style>" % CSS]
    for name, mc, mr, charts in plan:
        body.append("<h2>%s</h2>" % html.escape(name))
        body.append(render_sheet(wb[name], mc, mr, charts))
    open(out_html, "w", encoding="utf-8").write(
        "<!doctype html><meta charset='utf-8'>" + "\n".join(body))
    print("geschrieben:", out_html)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "Cashflow-Fruehwarnsystem.xlsx",
         sys.argv[2] if len(sys.argv) > 2 else "preview.html")
