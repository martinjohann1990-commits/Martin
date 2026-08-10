"""Einheitliches Farbschema und Zellstile fuer alle Blaetter der Arbeitsmappe.

Alle Formatierungen der Arbeitsmappe laufen ueber die hier definierten
benannten Zellstile (``NamedStyle``). Damit gibt es keine Ad-hoc-Formatierung
in den Blatt-Bausteinen und das Layout bleibt aenderbar an einer Stelle.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

FONT = "Arial"

# --- Farbpalette (auch von den Diagrammen und vom VBA-Setup verwendet) -------
DUNKELBLAU = "1F3864"
BLAU = "2E75B6"
HELLBLAU = "DEEBF7"
AKZENT = "ED7D31"
GRUEN = "70AD47"
ROT = "C00000"
GRAU = "595959"
HELLGRAU = "F2F2F2"
WEISS = "FFFFFF"
EINGABE_FILL = "FFF2CC"
EINGABE_FONT = "0000FF"  # Konvention: blaue Schrift = manuelle Eingabe
VERWEIS_FONT = "008000"  # gruene Schrift = Verweis auf ein anderes Blatt

# Reihenfolge der Diagrammfarben
SERIENFARBEN = [BLAU, AKZENT, GRUEN, DUNKELBLAU, "A5A5A5", "FFC000", "7030A0", "C55A11"]

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style(name: str, **kwargs) -> NamedStyle:
    style = NamedStyle(name=name)
    style.font = kwargs.get("font", Font(name=FONT, size=10, color=GRAU))
    if "fill" in kwargs:
        style.fill = kwargs["fill"]
    if "alignment" in kwargs:
        style.alignment = kwargs["alignment"]
    if kwargs.get("border", False):
        style.border = _BORDER
    if "number_format" in kwargs:
        style.number_format = kwargs["number_format"]
    return style


def styles() -> list[NamedStyle]:
    """Liefert alle Zellstile der Arbeitsmappe."""
    links = Alignment(horizontal="left", vertical="center")
    rechts = Alignment(horizontal="right", vertical="center")
    mitte = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return [
        _style(
            "LNP_Titel",
            font=Font(name=FONT, size=16, bold=True, color=DUNKELBLAU),
            alignment=links,
        ),
        _style(
            "LNP_Untertitel",
            font=Font(name=FONT, size=10, italic=True, color=GRAU),
            alignment=links,
        ),
        _style(
            "LNP_Abschnitt",
            font=Font(name=FONT, size=11, bold=True, color=DUNKELBLAU),
            alignment=links,
        ),
        _style(
            "LNP_TabKopf",
            font=Font(name=FONT, size=10, bold=True, color=WEISS),
            fill=PatternFill("solid", fgColor=DUNKELBLAU),
            alignment=mitte,
            border=True,
        ),
        _style("LNP_Text", alignment=links, border=True),
        _style("LNP_Label", font=Font(name=FONT, size=10, bold=True, color=DUNKELBLAU), alignment=links),
        _style(
            "LNP_Eingabe",
            font=Font(name=FONT, size=10, bold=True, color=EINGABE_FONT),
            fill=PatternFill("solid", fgColor=EINGABE_FILL),
            alignment=rechts,
            border=True,
            number_format="#,##0.00",
        ),
        _style(
            "LNP_EingabeText",
            font=Font(name=FONT, size=10, bold=True, color=EINGABE_FONT),
            fill=PatternFill("solid", fgColor=EINGABE_FILL),
            alignment=links,
            border=True,
        ),
        _style(
            "LNP_EingabeProzent",
            font=Font(name=FONT, size=10, bold=True, color=EINGABE_FONT),
            fill=PatternFill("solid", fgColor=EINGABE_FILL),
            alignment=rechts,
            border=True,
            number_format="0.0%",
        ),
        _style("LNP_Zahl", alignment=rechts, border=True, number_format="#,##0"),
        _style("LNP_Zahl2", alignment=rechts, border=True, number_format="#,##0.00"),
        _style("LNP_Prozent", alignment=rechts, border=True, number_format="0.0%"),
        _style("LNP_Score", alignment=rechts, border=True, number_format="0.000"),
        _style("LNP_Euro", alignment=rechts, border=True, number_format='#,##0.00 "€"'),
        _style("LNP_Datum", alignment=rechts, border=True, number_format="DD.MM.YYYY"),
        _style(
            "LNP_Verweis",
            font=Font(name=FONT, size=10, color=VERWEIS_FONT),
            alignment=rechts,
            border=True,
            number_format="#,##0.00",
        ),
        _style(
            "LNP_VerweisText",
            font=Font(name=FONT, size=10, color=VERWEIS_FONT),
            alignment=links,
            border=True,
        ),
        _style(
            "LNP_KPI_Wert",
            font=Font(name=FONT, size=22, bold=True, color=DUNKELBLAU),
            fill=PatternFill("solid", fgColor=HELLBLAU),
            alignment=Alignment(horizontal="center", vertical="center"),
            number_format="#,##0",
        ),
        _style(
            "LNP_KPI_WertText",
            font=Font(name=FONT, size=14, bold=True, color=DUNKELBLAU),
            fill=PatternFill("solid", fgColor=HELLBLAU),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        _style(
            "LNP_KPI_WertProzent",
            font=Font(name=FONT, size=22, bold=True, color=DUNKELBLAU),
            fill=PatternFill("solid", fgColor=HELLBLAU),
            alignment=Alignment(horizontal="center", vertical="center"),
            number_format="0.0%",
        ),
        _style(
            "LNP_KPI_Label",
            font=Font(name=FONT, size=9, bold=True, color=BLAU),
            fill=PatternFill("solid", fgColor=HELLBLAU),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        ),
        _style(
            "LNP_KPI_Fuss",
            font=Font(name=FONT, size=8, color=GRAU),
            fill=PatternFill("solid", fgColor=HELLBLAU),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        _style(
            "LNP_Ergebnis",
            font=Font(name=FONT, size=13, bold=True, color=WEISS),
            fill=PatternFill("solid", fgColor=GRUEN),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        _style(
            "LNP_Hinweis",
            font=Font(name=FONT, size=9, italic=True, color=GRAU),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        ),
        _style(
            "LNP_Warnung",
            font=Font(name=FONT, size=10, bold=True, color=ROT),
            alignment=links,
        ),
    ]


def register(wb) -> None:
    """Registriert alle Stile in der Arbeitsmappe."""
    vorhanden = set(wb.named_styles)
    for style in styles():
        if style.name not in vorhanden:
            wb.add_named_style(style)
