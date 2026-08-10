"""Milestone 1 acceptance check: the generated file opens cleanly and the
other non-negotiable technical constraints from spec section 2 hold -
.xlsx (never .xlsm), no macros, sheet protection set without a password,
and Method hidden. test_no_banned_functions / test_references_resolve /
test_empty_workbook_clean cover the rest of section 2."""

from pathlib import Path

import openpyxl
import pytest

from build.main import generate
from build.workbook import SHEET_MODULES, sheet_title
from tests._helpers import all_workbooks


@pytest.mark.parametrize("vertical,lang,wb", all_workbooks())
def test_sheet_order_matches_spec(vertical, lang, wb):
    from build.workbook import load_context
    ctx = load_context(vertical, lang)
    expected = [sheet_title(ctx, key) for key, _ in SHEET_MODULES]
    assert wb.sheetnames == expected


@pytest.mark.parametrize("vertical,lang,wb", all_workbooks())
def test_method_sheet_is_hidden(vertical, lang, wb):
    from build.workbook import load_context
    ctx = load_context(vertical, lang)
    method_title = sheet_title(ctx, "method")
    assert wb[method_title].sheet_state == "hidden"


@pytest.mark.parametrize("vertical,lang,wb", all_workbooks())
def test_every_sheet_protected_without_password(vertical, lang, wb):
    for ws in wb.worksheets:
        assert ws.protection.sheet is True, f"{ws.title} is not protected"
        assert not ws.protection.password, f"{ws.title} has a protection password set"


def test_generated_file_round_trips(tmp_path):
    out_path = generate("candles", "en", tmp_path)
    assert out_path.name == "Peak-Season-Planner-EN.xlsx"
    assert out_path.suffix == ".xlsx"

    reopened = openpyxl.load_workbook(out_path)
    assert reopened.sheetnames[0] == "Start"
    assert reopened.vba_archive is None
