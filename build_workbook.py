#!/usr/bin/env python3
"""Baut die Arbeitsmappe "Logistik_Netzwerkplanung.xlsm".

Aufruf:  python3 build_workbook.py [Zieldatei]

Der Build laeuft in der von der Aufgabenstellung vorgegebenen Reihenfolge:
DC_Stammdaten + Import-Blaetter -> Zielreichweite -> Simulation/Scoring ->
Dashboard/KPIs -> Szenarien + Export. Anschliessend wird das VBA-Projekt
erzeugt und die Datei als makrofaehige .xlsm geschrieben.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, BubbleChart, LineChart, Reference, Series
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).parent / "tools"))

import beispieldaten as bd  # noqa: E402
import design as dsg  # noqa: E402
import ovba  # noqa: E402
import xlsm  # noqa: E402

# --- feste Blattpositionen (an einer Stelle, damit Formeln konsistent sind) --
DATA_FIRST = 5  # erste Datenzeile der Import-Blaetter
DATA_LAST = 5004  # Ende des Formel-Fensters der Import-Blaetter
DC_FIRST, DC_LAST = 6, 11  # Datenzeilen DC_Stammdaten (6 DCs)
DC_MAX = 25  # bis hierhin reichen die Aggregat-Formeln (Platz fuer neue DCs)
REG_FIRST, REG_LAST = 6, 13  # Datenzeilen Regionstabelle (8 Regionen)
ZR_FIRST, ZR_LAST = 14, 18  # Datenzeilen Zielreichweite (5 Kategorien)
SIM_FIRST, SIM_LAST = 20, 25  # Datenzeilen der Berechnungstabelle Simulation
SIM_REG_FIRST, SIM_REG_LAST = 6, 13  # Regionsraster auf dem Blatt Simulation
SZ_FIRST, SZ_LAST = 7, 36  # Snapshot-Zeilen Szenarien
KPI_MONAT_FIRST, KPI_MONAT_LAST = 51, 62  # Monatsraster Dashboard (12 Monate)

HIST = "Import_Historie"
FC = "Import_Forecast"

# Spaltenbuchstaben der Import-Blaetter
SP_DATUM, SP_REGION, SP_KAT, SP_MENGE = "A", "C", "D", "E"
SP_UMSATZ, SP_VOLUMEN, SP_PAL, SP_PALAEQ = "F", "G", "H", "I"


def rng(sheet: str, col: str) -> str:
    """Absoluter Bereich einer Import-Spalte, z.B. Import_Historie!$I$5:$I$5004."""
    return f"{sheet}!${col}${DATA_FIRST}:${col}${DATA_LAST}"


# ---------------------------------------------------------------------------
# kleine Helfer
# ---------------------------------------------------------------------------


def put(ws, addr: str, value=None, style: str | None = None, fmt: str | None = None):
    cell = ws[addr]
    if value is not None:
        cell.value = value
    if style:
        cell.style = style
    if fmt:
        cell.number_format = fmt
    return cell


def merge(ws, ref: str, value=None, style: str | None = None):
    """Verbindet ``ref`` und formatiert alle Zellen des Bereichs einheitlich."""
    ws.merge_cells(ref)
    first = ref.split(":")[0]
    for row in ws[ref]:
        for cell in row:
            if style:
                cell.style = style
    if value is not None:
        ws[first].value = value
    return ws[first]


def kopfzeile(ws, row: int, spalten: list[str], first_col: int = 1, hoehe: int = 30):
    for i, text in enumerate(spalten):
        put(ws, f"{get_column_letter(first_col + i)}{row}", text, "LNP_TabKopf")
    ws.row_dimensions[row].height = hoehe


def breiten(ws, spec: dict[str, float]):
    for col, width in spec.items():
        ws.column_dimensions[col].width = width


def tabelle(ws, name: str, ref: str, stil: str = "TableStyleMedium2"):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=stil, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    return table


def titel(ws, text: str, untertitel: str | None = None):
    put(ws, "A1", text, "LNP_Titel")
    ws.row_dimensions[1].height = 24
    if untertitel:
        put(ws, "A2", untertitel, "LNP_Untertitel")
    ws.sheet_view.showGridLines = False


def stil_diagramm(chart, titel_text: str, hoehe: float = 7.5, breite: float = 14.0):
    chart.title = titel_text
    chart.height = hoehe
    chart.width = breite
    chart.style = 2
    for i, serie in enumerate(chart.series):
        farbe = dsg.SERIENFARBEN[i % len(dsg.SERIENFARBEN)]
        serie.graphicalProperties.solidFill = farbe
        serie.graphicalProperties.line.solidFill = farbe
    return chart


# ---------------------------------------------------------------------------
# 1. DC_Stammdaten
# ---------------------------------------------------------------------------

DC_SPALTEN = [
    "DC-Name",
    "Region/Land",
    "Latitude",
    "Longitude",
    "Kapazität (Paletten)",
    "Lagerkosten (€/Palette/Periode)",
    "Transportkosten-Basis (€/Palette)",
    "Transportkosten (€/km/Palette)",
    "Vorbelegung (Paletten)",
    "aktiv (Ja/Nein)",
]


def blatt_dc(wb):
    ws = wb.create_sheet("DC_Stammdaten")
    titel(
        ws,
        "Distributionszentren – Stammdaten",
        "Blaue Zellen sind Eingaben. Zeilen ausschließlich über die Schaltflächen "
        '"DC hinzufügen" / "DC entfernen" ändern, damit die Formeln der Simulation '
        "mitwachsen.",
    )
    breiten(
        ws,
        {"A": 18, "B": 14, "C": 11, "D": 11, "E": 18, "F": 22, "G": 22, "H": 22,
         "I": 18, "J": 14, "K": 3, "M": 14, "N": 14, "O": 11, "P": 11},
    )

    kopfzeile(ws, 5, DC_SPALTEN)
    for i, dc in enumerate(bd.DCS):
        row = DC_FIRST + i
        put(ws, f"A{row}", dc[0], "LNP_EingabeText")
        put(ws, f"B{row}", dc[1], "LNP_EingabeText")
        put(ws, f"C{row}", dc[2], "LNP_Eingabe", "0.0000")
        put(ws, f"D{row}", dc[3], "LNP_Eingabe", "0.0000")
        put(ws, f"E{row}", dc[4], "LNP_Eingabe", "#,##0")
        put(ws, f"F{row}", dc[5], "LNP_Eingabe", '#,##0.00 "€"')
        put(ws, f"G{row}", dc[6], "LNP_Eingabe", '#,##0.00 "€"')
        put(ws, f"H{row}", dc[7], "LNP_Eingabe", '#,##0.000 "€"')
        put(ws, f"I{row}", dc[8], "LNP_Eingabe", "#,##0")
        put(ws, f"J{row}", dc[9], "LNP_EingabeText")
    tabelle(ws, "tblDC", f"A5:J{DC_LAST}")

    dv = DataValidation(type="list", formula1='"Ja,Nein"', allow_blank=True)
    dv.add(f"J{DC_FIRST}:J{DC_MAX}")
    ws.add_data_validation(dv)

    # Absatzregionen: Referenztabelle fuer Distanzen und die Karte
    put(ws, "M3", "Absatzregionen (Kunden-Schwerpunkte)", "LNP_Abschnitt")
    kopfzeile(ws, 5, ["Region", "Land", "Latitude", "Longitude"], first_col=13)
    for i, reg in enumerate(bd.REGIONEN):
        row = REG_FIRST + i
        put(ws, f"M{row}", reg[0], "LNP_EingabeText")
        put(ws, f"N{row}", reg[1], "LNP_EingabeText")
        put(ws, f"O{row}", reg[2], "LNP_Eingabe", "0.0000")
        put(ws, f"P{row}", reg[3], "LNP_Eingabe", "0.0000")
    tabelle(ws, "tblRegionen", f"M5:P{REG_LAST}", "TableStyleMedium7")

    ws.freeze_panes = "A6"
    return ws


# ---------------------------------------------------------------------------
# 2. Import_Historie / Import_Forecast / Import_Mapping
# ---------------------------------------------------------------------------


def blatt_import(wb, name: str, ueberschrift: str, hinweis: str, zeilen: list[list],
                 tabellenname: str):
    ws = wb.create_sheet(name)
    titel(ws, ueberschrift, hinweis)
    breiten(
        ws,
        {"A": 13, "B": 22, "C": 13, "D": 19, "E": 14, "F": 15, "G": 17, "H": 12,
         "I": 19, "K": 60},
    )
    kopfzeile(ws, 4, bd.ZIEL_SPALTEN)

    for i, zeile in enumerate(zeilen):
        row = DATA_FIRST + i
        put(ws, f"A{row}", zeile[0], "LNP_Datum")
        put(ws, f"B{row}", zeile[1], "LNP_Text")
        put(ws, f"C{row}", zeile[2], "LNP_Text")
        put(ws, f"D{row}", zeile[3], "LNP_Text")
        put(ws, f"E{row}", zeile[4], "LNP_Zahl")
        put(ws, f"F{row}", zeile[5], "LNP_Euro")
        put(ws, f"G{row}", zeile[6], "LNP_Zahl2")
        put(ws, f"H{row}", zeile[7], "LNP_Zahl2")
        # Paletten-Äquivalent: gemeldete Paletten, sonst Volumen / m³ je Palette
        put(
            ws,
            f"I{row}",
            f"=IF($H{row}>0,$H{row},IFERROR($G{row}/Import_Mapping!$B$7,0))",
            "LNP_Zahl2",
        )
    letzte = DATA_FIRST + len(zeilen) - 1
    tabelle(ws, tabellenname, f"A4:I{letzte}")

    put(
        ws,
        "K2",
        f"Hinweis: Alle Auswertungen lesen die Zeilen {DATA_FIRST} bis {DATA_LAST}. "
        "Enthält die Quelle mehr Zeilen, muss dieses Fenster in den Formeln "
        "erweitert werden.",
        "LNP_Hinweis",
    )
    ws.freeze_panes = "A5"
    return ws


MAPPING_FELDER = [
    ("Datum", "Buchungsdatum", "Planperiode", "Belegdatum bzw. Planperiode"),
    ("Kunde", "Kunde_Name", "Kunde", "Kundenname oder -nummer"),
    ("Land/Region", "Land_Region", "Vertriebsregion", "muss zu 'Region' in DC_Stammdaten passen"),
    ("Produktkategorie", "Warengruppe", "Produktgruppe", "muss zu Zielreichweite passen"),
    ("Menge (Stück)", "Menge_Stk", "Planmenge_Stk", "Absatzmenge in Stück"),
    ("EUR-Umsatz", "Umsatz_EUR", "Planumsatz_EUR", "Umsatz in EUR"),
    ("Transportvolumen", "Volumen_m3", "Planvolumen_m3", "Volumen in m³"),
    ("Paletten", "Paletten_Anzahl", "Planpaletten", "gemeldete Paletten (optional)"),
]


def blatt_mapping(wb):
    ws = wb.create_sheet("Import_Mapping")
    titel(
        ws,
        "Import-Parameter und Spalten-Mapping",
        "Die Power-Query-Abfragen lesen diese Tabelle. Trägt die Quelldatei andere "
        "Spaltennamen, wird nur die Spalte B bzw. C angepasst – nicht die Abfrage.",
    )
    breiten(ws, {"A": 30, "B": 46, "C": 46, "D": 44})

    put(ws, "A4", "Quelldatei Historie (vollständiger Pfad)", "LNP_Label")
    put(ws, "B4", "", "LNP_EingabeText")
    put(ws, "A5", "Quelldatei Forecast (vollständiger Pfad)", "LNP_Label")
    put(ws, "B5", "", "LNP_EingabeText")
    put(ws, "A6", "CSV-Trennzeichen", "LNP_EingabeText")
    put(ws, "B6", ";", "LNP_EingabeText")
    put(ws, "A7", "m³ pro Palette (für Paletten-Äquivalent)", "LNP_Label")
    put(ws, "B7", 1.5, "LNP_Eingabe", "0.00")
    put(ws, "A8", "Name der Excel-Tabelle/des Blatts in der Quelldatei", "LNP_Label")
    put(ws, "B8", "", "LNP_EingabeText")
    put(
        ws,
        "D4",
        "Die Pfade werden von den Schaltflächen \"Datei importieren\" gefüllt. "
        "Bleibt der Pfad leer, arbeitet die Mappe mit den mitgelieferten "
        "Beispieldaten weiter.",
        "LNP_Hinweis",
    )
    put(
        ws,
        "D7",
        "Wird für Zeilen ohne gemeldete Paletten verwendet: "
        "Paletten-Äquivalent = Transportvolumen / m³ pro Palette.",
        "LNP_Hinweis",
    )

    # Die Spaltenköpfe werden vom M-Code der Power-Query-Abfragen namentlich
    # gelesen (Record.Field) – sie dürfen nicht umbenannt werden.
    kopfzeile(ws, 11, ["Zielfeld", "Quellspalte Historie",
                       "Quellspalte Forecast", "Hinweis"])
    for i, (ziel, qh, qf, hinweis) in enumerate(MAPPING_FELDER):
        row = 12 + i
        put(ws, f"A{row}", ziel, "LNP_Text")
        put(ws, f"B{row}", qh, "LNP_EingabeText")
        put(ws, f"C{row}", qf, "LNP_EingabeText")
        put(ws, f"D{row}", hinweis, "LNP_Text")
    tabelle(ws, "tblMapping", f"A11:D{11 + len(MAPPING_FELDER)}")

    put(
        ws,
        "A22",
        "Das Zielfeld \"Paletten-Äquivalent\" wird nicht gemappt, sondern berechnet.",
        "LNP_Hinweis",
    )
    return ws


# ---------------------------------------------------------------------------
# 3. Zielreichweite
# ---------------------------------------------------------------------------


def blatt_zielreichweite(wb):
    ws = wb.create_sheet("Zielreichweite")
    titel(
        ws,
        "Zielreichweite und Ziel-Bestand je Produktkategorie",
        "Ziel-Bestand = Ø Bedarf/Tag × Ziel-Reichweite. Die Bedarfsbasis ist "
        "umschaltbar zwischen Historie, Forecast und einem gewichteten Mix.",
    )
    breiten(ws, {"A": 26, "B": 20, "C": 22, "D": 22, "E": 22, "F": 22, "H": 58})

    put(ws, "A4", "Steuerung der Bedarfsbasis", "LNP_Abschnitt")
    put(ws, "A5", "Bedarfsbasis", "LNP_Label")
    put(ws, "B5", "Mix", "LNP_EingabeText")
    put(ws, "A6", "Gewicht Historie (nur bei Mix)", "LNP_Label")
    put(ws, "B6", 0.6, "LNP_EingabeProzent")
    put(ws, "A7", "Gewicht Forecast (nur bei Mix)", "LNP_Label")
    put(ws, "B7", "=1-$B$6", "LNP_Prozent")
    put(ws, "A8", "Tage im Historie-Zeitraum", "LNP_Label")
    put(
        ws,
        "B8",
        f"=IF(COUNT({rng(HIST, SP_DATUM)})=0,0,"
        f"MAX({rng(HIST, SP_DATUM)})-MIN({rng(HIST, SP_DATUM)})+1)",
        "LNP_Zahl",
    )
    put(ws, "A9", "Tage im Forecast-Zeitraum", "LNP_Label")
    put(
        ws,
        "B9",
        f"=IF(COUNT({rng(FC, SP_DATUM)})=0,0,"
        f"MAX({rng(FC, SP_DATUM)})-MIN({rng(FC, SP_DATUM)})+1)",
        "LNP_Zahl",
    )
    put(ws, "A10", "Bewegungszeilen Historie / Forecast", "LNP_Label")
    put(ws, "B10", f"=COUNT({rng(HIST, SP_DATUM)})", "LNP_Zahl")
    put(ws, "C10", f"=COUNT({rng(FC, SP_DATUM)})", "LNP_Zahl")

    dv = DataValidation(type="list", formula1='"Historie,Forecast,Mix"', allow_blank=False)
    dv.add("B5")
    ws.add_data_validation(dv)
    dv2 = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv2.error = "Bitte einen Wert zwischen 0 % und 100 % eingeben."
    dv2.add("B6")
    ws.add_data_validation(dv2)

    put(
        ws,
        "H5",
        "Ø Bedarf/Tag = Summe Paletten-Äquivalent der Kategorie geteilt durch die "
        "Anzahl Tage des jeweiligen Zeitraums (Annahme: durchgehender Bedarf über "
        "den gesamten Zeitraum).",
        "LNP_Hinweis",
    )

    put(ws, "A12", "Ziel-Reichweite je Produktkategorie", "LNP_Abschnitt")
    kopfzeile(
        ws,
        13,
        [
            "Produktkategorie",
            "Ziel-Reichweite (Tage)",
            "Ø Bedarf/Tag Historie (Pal.)",
            "Ø Bedarf/Tag Forecast (Pal.)",
            "Ø Bedarf/Tag Basis (Pal.)",
            "Ziel-Bestand (Paletten)",
        ],
    )
    for i, kat in enumerate(bd.KATEGORIEN):
        row = ZR_FIRST + i
        put(ws, f"A{row}", kat[0], "LNP_EingabeText")
        put(ws, f"B{row}", kat[1], "LNP_Eingabe", "#,##0")
        put(
            ws,
            f"C{row}",
            f"=IFERROR(SUMIFS({rng(HIST, SP_PALAEQ)},{rng(HIST, SP_KAT)},$A{row})/$B$8,0)",
            "LNP_Zahl2",
        )
        put(
            ws,
            f"D{row}",
            f"=IFERROR(SUMIFS({rng(FC, SP_PALAEQ)},{rng(FC, SP_KAT)},$A{row})/$B$9,0)",
            "LNP_Zahl2",
        )
        put(
            ws,
            f"E{row}",
            f'=IF($B$5="Historie",$C{row},IF($B$5="Forecast",$D{row},'
            f"$B$6*$C{row}+$B$7*$D{row}))",
            "LNP_Zahl2",
        )
        put(ws, f"F{row}", f"=ROUND($E{row}*$B{row},0)", "LNP_Zahl")
    tabelle(ws, "tblZielreichweite", f"A13:F{ZR_LAST}")

    put(ws, f"A{ZR_LAST + 1}", "Summe", "LNP_Label")
    put(ws, f"E{ZR_LAST + 1}", f"=SUM($E${ZR_FIRST}:$E${ZR_LAST})", "LNP_Zahl2")
    put(ws, f"F{ZR_LAST + 1}", f"=SUM($F${ZR_FIRST}:$F${ZR_LAST})", "LNP_Zahl")

    ws.conditional_formatting.add(
        f"F{ZR_FIRST}:F{ZR_LAST}",
        ColorScaleRule(
            start_type="min", start_color=dsg.HELLBLAU,
            end_type="max", end_color=dsg.BLAU,
        ),
    )
    return ws


# ---------------------------------------------------------------------------
# 4. Simulation
# ---------------------------------------------------------------------------

SIM_SPALTEN = [
    "DC-Name",
    "aktiv",
    "Kapazität (Pal.)",
    "Vorbelegung (Pal.)",
    "Ziel-Bestand Kategorie (Pal.)",
    "Auslastung nach Zuordnung",
    "Kapazitäts-Score (roh)",
    "Distanz zum Bedarfs-schwerpunkt (km)",
    "Transportkosten (€/Pal.)",
    "Transport-Score (roh = –Kosten)",
    "erreichbarer Bestand (Pal.)",
    "Reichweiten-Score (roh)",
    "Teil-Score Kapazität (0–1)",
    "Teil-Score Transport (0–1)",
    "Teil-Score Reichweite (0–1)",
    "Gesamt-Score (gewichtet)",
    "Sortier-schlüssel",
    "Rang",
    "Score für Splitting",
    "empfohlener Anteil",
    "Empfehlung",
]


def blatt_simulation(wb):
    ws = wb.create_sheet("Simulation")
    titel(
        ws,
        "Simulation – Zuordnung Produktkategorie zu Distributionszentrum",
        "Jeder Teil-Score steht in einer eigenen Spalte und ist damit "
        "nachvollziehbar. Der Gesamt-Score ist die gewichtete Summe der drei "
        "Teil-Scores.",
    )
    breiten(
        ws,
        {"A": 20, "B": 16, "C": 30, "D": 14, "E": 14, "F": 16, "G": 16, "H": 14,
         "I": 14, "J": 15, "K": 14, "L": 15, "M": 14, "N": 14, "O": 14, "P": 15,
         "Q": 13, "R": 8, "S": 14, "T": 14, "U": 13},
    )

    # --- Eingaben ---------------------------------------------------------
    put(ws, "A3", "Eingaben", "LNP_Abschnitt")
    put(ws, "A4", "Produktkategorie", "LNP_Label")
    put(ws, "B4", bd.KATEGORIEN[2][0], "LNP_EingabeText")
    put(ws, "A5", "Bedarfsbasis (aus Zielreichweite)", "LNP_Label")
    put(ws, "B5", "=Zielreichweite!$B$5", "LNP_VerweisText")
    put(ws, "A6", "Ziel-Reichweite (Tage)", "LNP_Label")
    put(
        ws,
        "B6",
        f"=IFERROR(INDEX(Zielreichweite!$B${ZR_FIRST}:$B${ZR_LAST},"
        f"MATCH($B$4,Zielreichweite!$A${ZR_FIRST}:$A${ZR_LAST},0)),0)",
        "LNP_Verweis",
        "#,##0",
    )
    put(ws, "A7", "Ø Bedarf/Tag (Paletten)", "LNP_Label")
    put(
        ws,
        "B7",
        f"=IFERROR(INDEX(Zielreichweite!$E${ZR_FIRST}:$E${ZR_LAST},"
        f"MATCH($B$4,Zielreichweite!$A${ZR_FIRST}:$A${ZR_LAST},0)),0)",
        "LNP_Verweis",
        "#,##0.00",
    )
    put(ws, "A8", "Ziel-Bestand (Paletten)", "LNP_Label")
    put(
        ws,
        "B8",
        f"=IFERROR(INDEX(Zielreichweite!$F${ZR_FIRST}:$F${ZR_LAST},"
        f"MATCH($B$4,Zielreichweite!$A${ZR_FIRST}:$A${ZR_LAST},0)),0)",
        "LNP_Verweis",
        "#,##0",
    )

    put(ws, "A10", "Gewichtung Kapazitätsauslastung", "LNP_Label")
    put(ws, "B10", 0.40, "LNP_EingabeProzent")
    put(ws, "A11", "Gewichtung Transportkosten/Distanz", "LNP_Label")
    put(ws, "B11", 0.35, "LNP_EingabeProzent")
    put(ws, "A12", "Gewichtung Zielreichweite-Einhaltung", "LNP_Label")
    put(ws, "B12", 0.25, "LNP_EingabeProzent")
    put(ws, "A13", "Summe der Gewichtungen", "LNP_Label")
    put(ws, "B13", "=SUM($B$10:$B$12)", "LNP_Prozent")
    put(
        ws,
        "C13",
        '=IF(ROUND($B$13,4)=1,"OK – Summe ergibt 100 %",'
        '"FEHLER – die drei Gewichtungen müssen zusammen 100 % ergeben")',
        "LNP_Warnung",
    )

    put(ws, "A15", "Zuordnungsmodus", "LNP_Label")
    put(ws, "B15", "Splitting auf mehrere DCs", "LNP_EingabeText")
    put(ws, "A16", "max. Anzahl DCs bei Splitting", "LNP_Label")
    put(ws, "B16", 3, "LNP_Eingabe", "0")

    dv_kat = DataValidation(
        type="list",
        formula1=f"=Zielreichweite!$A${ZR_FIRST}:$A${ZR_LAST}",
        allow_blank=False,
    )
    dv_kat.add("B4")
    ws.add_data_validation(dv_kat)
    dv_gew = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv_gew.error = "Zulässig sind Werte von 0 % bis 100 %."
    dv_gew.errorTitle = "Ungültige Gewichtung"
    dv_gew.add("B10:B12")
    ws.add_data_validation(dv_gew)
    dv_modus = DataValidation(
        type="list",
        formula1='"Alleinzuordnung,Splitting auf mehrere DCs"',
        allow_blank=False,
    )
    dv_modus.add("B15")
    ws.add_data_validation(dv_modus)
    dv_anz = DataValidation(type="whole", operator="between", formula1=1, formula2=DC_MAX)
    dv_anz.add("B16")
    ws.add_data_validation(dv_anz)

    ws.conditional_formatting.add(
        "B13",
        FormulaRule(
            formula=["ROUND($B$13,4)<>1"],
            fill=PatternFill("solid", bgColor="FFC7CE"),
            font=Font(name=dsg.FONT, bold=True, color="9C0006"),
        ),
    )

    # --- Bedarfsschwerpunkt der gewaehlten Kategorie ----------------------
    put(ws, "F3", "Bedarfsschwerpunkt der gewählten Kategorie", "LNP_Abschnitt")
    kopfzeile(
        ws,
        5,
        [
            "Region",
            "Bedarf/Tag Historie (Pal.)",
            "Bedarf/Tag Forecast (Pal.)",
            "Bedarf/Tag Basis (Pal.)",
            "Latitude",
            "Longitude",
        ],
        first_col=6,
    )
    for i in range(SIM_REG_LAST - SIM_REG_FIRST + 1):
        row = SIM_REG_FIRST + i
        dc_row = REG_FIRST + i
        put(ws, f"F{row}", f"=DC_Stammdaten!$M{dc_row}", "LNP_VerweisText")
        put(
            ws,
            f"G{row}",
            f"=IFERROR(SUMIFS({rng(HIST, SP_PALAEQ)},{rng(HIST, SP_KAT)},$B$4,"
            f"{rng(HIST, SP_REGION)},$F{row})/Zielreichweite!$B$8,0)",
            "LNP_Zahl2",
        )
        put(
            ws,
            f"H{row}",
            f"=IFERROR(SUMIFS({rng(FC, SP_PALAEQ)},{rng(FC, SP_KAT)},$B$4,"
            f"{rng(FC, SP_REGION)},$F{row})/Zielreichweite!$B$9,0)",
            "LNP_Zahl2",
        )
        put(
            ws,
            f"I{row}",
            f'=IF(Zielreichweite!$B$5="Historie",$G{row},'
            f'IF(Zielreichweite!$B$5="Forecast",$H{row},'
            f"Zielreichweite!$B$6*$G{row}+Zielreichweite!$B$7*$H{row}))",
            "LNP_Zahl2",
        )
        put(ws, f"J{row}", f"=DC_Stammdaten!$O{dc_row}", "LNP_Verweis", "0.0000")
        put(ws, f"K{row}", f"=DC_Stammdaten!$P{dc_row}", "LNP_Verweis", "0.0000")

    schwer = SIM_REG_LAST + 2  # Zeile 15
    put(ws, f"F{schwer}", "Summe / Schwerpunkt", "LNP_Label")
    put(ws, f"I{schwer}", f"=SUM($I${SIM_REG_FIRST}:$I${SIM_REG_LAST})", "LNP_Zahl2")
    put(
        ws,
        f"J{schwer}",
        f"=IFERROR(SUMPRODUCT($I${SIM_REG_FIRST}:$I${SIM_REG_LAST},"
        f"$J${SIM_REG_FIRST}:$J${SIM_REG_LAST})/$I${schwer},0)",
        "LNP_Zahl2",
        "0.0000",
    )
    put(
        ws,
        f"K{schwer}",
        f"=IFERROR(SUMPRODUCT($I${SIM_REG_FIRST}:$I${SIM_REG_LAST},"
        f"$K${SIM_REG_FIRST}:$K${SIM_REG_LAST})/$I${schwer},0)",
        "LNP_Zahl2",
        "0.0000",
    )

    # --- Berechnungstabelle je DC ----------------------------------------
    put(ws, "A18", "Bewertung je Distributionszentrum", "LNP_Abschnitt")
    kopfzeile(ws, 19, SIM_SPALTEN, hoehe=46)

    for i in range(SIM_LAST - SIM_FIRST + 1):
        r = SIM_FIRST + i
        d = DC_FIRST + i  # zugehoerige Zeile in DC_Stammdaten
        aktiv = f'OR($A{r}="",$B{r}<>"Ja")'  # Wächter: leere oder inaktive DCs

        put(ws, f"A{r}", f'=IF(DC_Stammdaten!$A{d}="","",DC_Stammdaten!$A{d})',
            "LNP_VerweisText")
        put(ws, f"B{r}", f'=IF($A{r}="","",DC_Stammdaten!$J{d})', "LNP_VerweisText")
        put(ws, f"C{r}", f'=IF($A{r}="","",DC_Stammdaten!$E{d})', "LNP_Verweis", "#,##0")
        put(ws, f"D{r}", f'=IF($A{r}="","",DC_Stammdaten!$I{d})', "LNP_Verweis", "#,##0")
        put(ws, f"E{r}", f'=IF($A{r}="","",$B$8)', "LNP_Verweis", "#,##0")
        put(ws, f"F{r}", f'=IF({aktiv},"",IFERROR(($D{r}+$E{r})/$C{r},0))',
            "LNP_Prozent")
        put(ws, f"G{r}", f'=IF({aktiv},"",1-$F{r})', "LNP_Score")
        # Orthodrome (Haversine über ACOS) zwischen DC und Bedarfsschwerpunkt
        put(
            ws,
            f"H{r}",
            f'=IF({aktiv},"",ROUND(6371*ACOS(MIN(1,'
            f"COS(RADIANS(DC_Stammdaten!$C{d}))*COS(RADIANS($J${schwer}))*"
            f"COS(RADIANS($K${schwer})-RADIANS(DC_Stammdaten!$D{d}))+"
            f"SIN(RADIANS(DC_Stammdaten!$C{d}))*SIN(RADIANS($J${schwer})))),0))",
            "LNP_Zahl",
        )
        put(
            ws,
            f"I{r}",
            f'=IF({aktiv},"",DC_Stammdaten!$G{d}+DC_Stammdaten!$H{d}*$H{r})',
            "LNP_Euro",
        )
        put(ws, f"J{r}", f'=IF({aktiv},"",-$I{r})', "LNP_Zahl2")
        put(ws, f"K{r}", f'=IF({aktiv},"",MIN($E{r},MAX(0,$C{r}-$D{r})))',
            "LNP_Zahl", "#,##0")
        put(
            ws,
            f"L{r}",
            f'=IF({aktiv},"",IFERROR(1-ABS($E{r}-$K{r})/$E{r},0))',
            "LNP_Score",
        )
        # Min-Max-Normierung der Rohwerte auf 0..1; sind alle Werte gleich,
        # liefert die Division einen Fehler und alle DCs bekommen 1.
        for ziel, quelle in (("M", "G"), ("N", "J"), ("O", "L")):
            bereich = f"${quelle}${SIM_FIRST}:${quelle}${SIM_LAST}"
            put(
                ws,
                f"{ziel}{r}",
                f'=IF({aktiv},"",IFERROR((${quelle}{r}-MIN({bereich}))'
                f"/(MAX({bereich})-MIN({bereich})),1))",
                "LNP_Score",
            )
        put(
            ws,
            f"P{r}",
            f'=IF({aktiv},"",$B$10*$M{r}+$B$11*$N{r}+$B$12*$O{r})',
            "LNP_Score",
        )
        put(ws, f"Q{r}", f"=IF({aktiv},-1,$P{r}-ROW()/1000000)", "LNP_Score", "0.000000")
        put(
            ws,
            f"R{r}",
            f'=IF({aktiv},"",SUMPRODUCT(($Q${SIM_FIRST}:$Q${SIM_LAST}>$Q{r})*1)+1)',
            "LNP_Zahl",
        )
        put(ws, f"S{r}", f"=IF({aktiv},0,IF($R{r}<=$B$16,MAX(0,$P{r}),0))", "LNP_Score")
        put(
            ws,
            f"T{r}",
            f'=IF({aktiv},"",IF($B$15="Alleinzuordnung",IF($R{r}=1,1,0),'
            f"IFERROR($S{r}/SUM($S${SIM_FIRST}:$S${SIM_LAST}),0)))",
            "LNP_Prozent",
        )
        put(
            ws,
            f"U{r}",
            f'=IF({aktiv},"",IF($B$15="Alleinzuordnung",IF($R{r}=1,"JA","nein"),'
            f'IF($T{r}>0,"anteilig","nein")))',
            "LNP_Text",
        )

    for spalte in ("M", "N", "O", "P"):
        ws.conditional_formatting.add(
            f"{spalte}{SIM_FIRST}:{spalte}{SIM_LAST}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=0.5, mid_color="FFEB84",
                end_type="num", end_value=1, end_color="63BE7B",
            ),
        )
    ws.conditional_formatting.add(
        f"T{SIM_FIRST}:T{SIM_LAST}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=dsg.BLAU),
    )
    ws.conditional_formatting.add(
        f"F{SIM_FIRST}:F{SIM_LAST}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="63BE7B",
            mid_type="num", mid_value=0.9, mid_color="FFEB84",
            end_type="num", end_value=1.2, end_color="F8696B",
        ),
    )

    # --- Ergebnis ---------------------------------------------------------
    erg = SIM_LAST + 2  # Zeile 27
    put(ws, f"A{erg}", "Ergebnis", "LNP_Abschnitt")
    merge(
        ws,
        f"A{erg + 1}:U{erg + 1}",
        f'=IF($B$15="Alleinzuordnung","Empfehlung: "&$B${erg + 3}&"   (Gesamt-Score "'
        f'&TEXT($D${erg + 3},"0.000")&")","Empfehlung Splitting: "&$B${erg + 3}&" "'
        f'&TEXT($C${erg + 3},"0 %")'
        f'&IF($C${erg + 4}=0,""," | "&$B${erg + 4}&" "&TEXT($C${erg + 4},"0 %"))'
        f'&IF($C${erg + 5}=0,""," | "&$B${erg + 5}&" "&TEXT($C${erg + 5},"0 %")))',
        "LNP_Ergebnis",
    )
    ws.row_dimensions[erg + 1].height = 30

    put(ws, f"A{erg + 2}", "Rangfolge (Hilfszeilen für die Ergebniszeile)", "LNP_Hinweis")
    for k in range(3):
        r = erg + 3 + k
        put(ws, f"A{r}", f"Rang {k + 1}", "LNP_Label")
        put(
            ws,
            f"B{r}",
            f'=IFERROR(INDEX($A${SIM_FIRST}:$A${SIM_LAST},'
            f"MATCH({k + 1},$R${SIM_FIRST}:$R${SIM_LAST},0)),\"\")",
            "LNP_Text",
        )
        put(
            ws,
            f"C{r}",
            f"=IFERROR(INDEX($T${SIM_FIRST}:$T${SIM_LAST},"
            f"MATCH({k + 1},$R${SIM_FIRST}:$R${SIM_LAST},0)),0)",
            "LNP_Prozent",
        )
        put(
            ws,
            f"D{r}",
            f"=IFERROR(INDEX($P${SIM_FIRST}:$P${SIM_LAST},"
            f"MATCH({k + 1},$R${SIM_FIRST}:$R${SIM_LAST},0)),0)",
            "LNP_Score",
        )
    ws.freeze_panes = "A20"
    return ws


# ---------------------------------------------------------------------------
# 5. Szenarien
# ---------------------------------------------------------------------------

SZ_SPALTEN = [
    "Szenarioname",
    "Zeitstempel",
    "Produktkategorie",
    "Bedarfsbasis",
    "Gewicht Kapazität",
    "Gewicht Transport",
    "Gewicht Reichweite",
    "Zuordnungsmodus",
    "Ziel-Reichweite (Tage)",
    "Ø Bedarf/Tag (Pal.)",
    "Ziel-Bestand (Pal.)",
    "Empfehlung (DC)",
    "bester Score",
]


def blatt_szenarien(wb):
    ws = wb.create_sheet("Szenarien")
    titel(
        ws,
        "Szenario-Vergleich",
        'Die Schaltfläche "Szenario speichern" schreibt den aktuellen Stand als '
        "Werte-Snapshot (ohne Formeln) in die nächste freie Zeile.",
    )
    n_dc = DC_LAST - DC_FIRST + 1
    spalten = SZ_SPALTEN + [f"Score DC {i + 1}" for i in range(n_dc)] + [
        f"Anteil DC {i + 1}" for i in range(n_dc)
    ]
    kopfzeile(ws, 6, spalten, hoehe=42)
    breiten(ws, {"A": 24, "B": 17, "C": 18, "D": 12, "E": 12, "F": 12, "G": 12,
                 "H": 22, "I": 13, "J": 13, "K": 13, "L": 16, "M": 11})
    for i in range(len(spalten) - 13):
        ws.column_dimensions[get_column_letter(14 + i)].width = 13

    put(
        ws,
        "A4",
        "Die Spaltenköpfe der Score- und Anteilsspalten werden beim Speichern "
        "automatisch mit den aktuellen DC-Namen beschriftet.",
        "LNP_Hinweis",
    )

    erste_score = 14  # Spalte N
    letzte_score = erste_score + n_dc - 1
    ws.conditional_formatting.add(
        f"{get_column_letter(erste_score)}{SZ_FIRST}:"
        f"{get_column_letter(letzte_score)}{SZ_LAST}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=0.5, mid_color="FFEB84",
            end_type="num", end_value=1, end_color="63BE7B",
        ),
    )

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    daten = Reference(ws, min_col=erste_score, max_col=letzte_score,
                      min_row=6, max_row=SZ_LAST)
    chart.add_data(daten, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=SZ_FIRST, max_row=SZ_LAST))
    stil_diagramm(chart, "Gesamt-Score je DC über alle Szenarien", hoehe=9, breite=24)
    chart.y_axis.title = "Gesamt-Score"
    chart.x_axis.title = "Szenario"
    ws.add_chart(chart, f"A{SZ_LAST + 3}")

    ws.freeze_panes = "A7"
    return ws


# ---------------------------------------------------------------------------
# 6. Karte
# ---------------------------------------------------------------------------


def blatt_karte(wb):
    ws = wb.create_sheet("Karte")
    titel(
        ws,
        "Karte (Näherung über XY-Streudiagramm)",
        "Native interaktive Karten sind in Excel eingeschränkt. Das Blasendiagramm "
        "zeigt Längengrad (X) gegen Breitengrad (Y); die Blasengröße entspricht der "
        "Kapazität bzw. dem Bedarf.",
    )
    breiten(ws, {"A": 20, "B": 12, "C": 12, "D": 18, "E": 3, "F": 14, "G": 12,
                 "H": 12, "I": 20, "K": 62})

    put(ws, "A5", "Distributionszentren", "LNP_Abschnitt")
    kopfzeile(ws, 6, ["DC-Name", "Longitude (X)", "Latitude (Y)", "Kapazität (Pal.)"])
    for i in range(DC_LAST - DC_FIRST + 1):
        r, d = 7 + i, DC_FIRST + i
        put(ws, f"A{r}", f'=IF(DC_Stammdaten!$A{d}="","",DC_Stammdaten!$A{d})',
            "LNP_VerweisText")
        put(ws, f"B{r}", f'=IF($A{r}="","",DC_Stammdaten!$D{d})', "LNP_Verweis", "0.0000")
        put(ws, f"C{r}", f'=IF($A{r}="","",DC_Stammdaten!$C{d})', "LNP_Verweis", "0.0000")
        put(ws, f"D{r}", f'=IF($A{r}="","",DC_Stammdaten!$E{d})', "LNP_Verweis", "#,##0")
    dc_letzte = 7 + (DC_LAST - DC_FIRST)

    put(ws, "F5", "Absatzregionen (gewählte Kategorie)", "LNP_Abschnitt")
    kopfzeile(ws, 6, ["Region", "Longitude (X)", "Latitude (Y)", "Bedarf/Tag (Pal.)"],
              first_col=6)
    for i in range(REG_LAST - REG_FIRST + 1):
        r = 7 + i
        sim_row = SIM_REG_FIRST + i
        put(ws, f"F{r}", f"=Simulation!$F{sim_row}", "LNP_VerweisText")
        put(ws, f"G{r}", f"=Simulation!$K{sim_row}", "LNP_Verweis", "0.0000")
        put(ws, f"H{r}", f"=Simulation!$J{sim_row}", "LNP_Verweis", "0.0000")
        put(ws, f"I{r}", f"=Simulation!$I{sim_row}", "LNP_Verweis", "#,##0.00")
    reg_letzte = 7 + (REG_LAST - REG_FIRST)

    chart = BubbleChart()
    chart.style = 18
    # Series(values=Y, xvalues=X, zvalues=Blasengroesse)
    serie_dc = Series(
        Reference(ws, min_col=3, min_row=7, max_row=dc_letzte),
        xvalues=Reference(ws, min_col=2, min_row=7, max_row=dc_letzte),
        zvalues=Reference(ws, min_col=4, min_row=7, max_row=dc_letzte),
        title="Distributionszentren",
    )
    serie_reg = Series(
        Reference(ws, min_col=8, min_row=7, max_row=reg_letzte),
        xvalues=Reference(ws, min_col=7, min_row=7, max_row=reg_letzte),
        zvalues=Reference(ws, min_col=9, min_row=7, max_row=reg_letzte),
        title="Absatzregionen",
    )
    chart.series.append(serie_dc)
    chart.series.append(serie_reg)
    stil_diagramm(chart, "Standorte und Bedarfsschwerpunkte", hoehe=12, breite=22)
    chart.x_axis.title = "Longitude"
    chart.y_axis.title = "Latitude"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    ws.add_chart(chart, "A17")

    put(
        ws,
        "K6",
        "Punkte beschriften: Diagramm anklicken → Datenbeschriftungen → "
        '"Werte aus Zellen" → Bereich A7:A' + str(dc_letzte) + " bzw. F7:F"
        + str(reg_letzte) + " wählen.",
        "LNP_Hinweis",
    )
    put(
        ws,
        "K9",
        "Zusatzoption: Für eine echte Karte die Spalten Region und Bedarf markieren "
        "und Einfügen → Karten → Kartendiagramm wählen (benötigt Excel 2019/365 mit "
        "Online-Verbindung). 3D-Karten (Power Map) über Einfügen → 3D-Karte.",
        "LNP_Hinweis",
    )
    return ws


# ---------------------------------------------------------------------------
# 7. Dashboard
# ---------------------------------------------------------------------------

KPI_KACHELN = [
    # Spaltenblock, Zeilenblock, Label, Formel, Stil, Fussnote, Sparkline-Spalte
    ("B", 7, "Bewegungszeilen Historie", "=Zielreichweite!$B$10", "LNP_KPI_Wert",
     "Zeilen im Import", "F"),
    ("F", 7, "Paletten-Äquivalent Historie", f"=SUM({rng(HIST, SP_PALAEQ)})",
     "LNP_KPI_Wert", "Summe über den Zeitraum", "C"),
    ("J", 7, "Umsatz Historie", f"=SUM({rng(HIST, SP_UMSATZ)})", "LNP_KPI_Wert",
     "EUR über den Zeitraum", "D"),
    ("B", 11, "Ø Bedarf/Tag alle Kategorien",
     f"=SUM(Zielreichweite!$E${ZR_FIRST}:$E${ZR_LAST})", "LNP_KPI_Wert",
     "Paletten pro Tag", "E"),
    ("F", 11, "Ziel-Bestand gesamt",
     f"=SUM(Zielreichweite!$F${ZR_FIRST}:$F${ZR_LAST})", "LNP_KPI_Wert",
     "Paletten laut Zielreichweite", "E"),
    ("J", 11, "Ø Auslastung aktive DCs",
     f'=IFERROR(SUMIFS(DC_Stammdaten!$I${DC_FIRST}:$I${DC_MAX},'
     f'DC_Stammdaten!$J${DC_FIRST}:$J${DC_MAX},"Ja")/'
     f'SUMIFS(DC_Stammdaten!$E${DC_FIRST}:$E${DC_MAX},'
     f'DC_Stammdaten!$J${DC_FIRST}:$J${DC_MAX},"Ja"),0)',
     "LNP_KPI_WertProzent", "Vorbelegung / Kapazität", "C"),
]


def blatt_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    titel(
        ws,
        "Logistik-Netzwerkplanung – Start",
        "Entscheidungshilfe: In welchem Distributionszentrum wird welche "
        "Produktkategorie gelagert?",
    )
    breiten(
        ws,
        {"A": 3, "B": 15, "C": 15, "D": 15, "E": 3, "F": 15, "G": 15, "H": 15,
         "I": 3, "J": 15, "K": 15, "L": 15, "M": 3},
    )
    put(ws, "B4", "Navigation (die Schaltflächen legt das Makro beim Öffnen an)",
        "LNP_Hinweis")
    ws.row_dimensions[5].height = 30

    for spalte, zeile, label, formel, stil, fuss, _spark in KPI_KACHELN:
        c1 = spalte
        c3 = get_column_letter(ws[f"{spalte}1"].column + 2)
        merge(ws, f"{c1}{zeile}:{c3}{zeile}", label, "LNP_KPI_Label")
        merge(ws, f"{c1}{zeile + 1}:{c3}{zeile + 1}", formel, stil)
        ws.row_dimensions[zeile + 1].height = 34
        merge(ws, f"{c1}{zeile + 2}:{c3}{zeile + 2}", None, "LNP_KPI_Fuss")
        put(ws, f"{c1}{zeile + 2}", fuss)
        ws.row_dimensions[zeile + 2].height = 22

    # Ergebnisband
    merge(
        ws,
        "B15:L15",
        '="Aktuelle Empfehlung für "&Simulation!$B$4&": "&Simulation!$B$30'
        '&"   |   Gesamt-Score "&TEXT(Simulation!$D$30,"0.000")'
        '&"   |   Modus: "&Simulation!$B$15',
        "LNP_Ergebnis",
    )
    ws.row_dimensions[15].height = 28

    # --- Diagramme --------------------------------------------------------
    ch1 = BarChart()
    ch1.type = "col"
    ch1.add_data(
        Reference(wb["DC_Stammdaten"], min_col=5, min_row=5, max_row=DC_LAST),
        titles_from_data=True,
    )
    ch1.add_data(
        Reference(wb["DC_Stammdaten"], min_col=9, min_row=5, max_row=DC_LAST),
        titles_from_data=True,
    )
    ch1.set_categories(
        Reference(wb["DC_Stammdaten"], min_col=1, min_row=DC_FIRST, max_row=DC_LAST)
    )
    stil_diagramm(ch1, "Kapazität und Vorbelegung je DC (Paletten)")
    ws.add_chart(ch1, "B17")

    ch2 = BarChart()
    ch2.type = "col"
    ch2.add_data(
        Reference(wb["Zielreichweite"], min_col=6, min_row=13, max_row=ZR_LAST),
        titles_from_data=True,
    )
    ch2.set_categories(
        Reference(wb["Zielreichweite"], min_col=1, min_row=ZR_FIRST, max_row=ZR_LAST)
    )
    stil_diagramm(ch2, "Ziel-Bestand je Produktkategorie (Paletten)")
    ws.add_chart(ch2, "H17")

    ch3 = BarChart()
    ch3.type = "bar"
    ch3.add_data(
        Reference(wb["Simulation"], min_col=9, min_row=5, max_row=SIM_REG_LAST),
        titles_from_data=True,
    )
    ch3.set_categories(
        Reference(wb["Simulation"], min_col=6, min_row=SIM_REG_FIRST,
                  max_row=SIM_REG_LAST)
    )
    stil_diagramm(ch3, "Ø Bedarf/Tag je Region (gewählte Kategorie)")
    ws.add_chart(ch3, "B33")

    ch4 = LineChart()
    ch4.add_data(
        Reference(ws, min_col=3, min_row=KPI_MONAT_FIRST - 1, max_row=KPI_MONAT_LAST),
        titles_from_data=True,
    )
    ch4.set_categories(
        Reference(ws, min_col=1, min_row=KPI_MONAT_FIRST, max_row=KPI_MONAT_LAST)
    )
    stil_diagramm(ch4, "Paletten-Äquivalent je Monat (Historie)")
    ch4.series[0].marker = Marker(symbol="circle", size=5)
    ws.add_chart(ch4, "H33")

    # --- Datenbasis fuer die Sparklines ----------------------------------
    put(ws, f"A{KPI_MONAT_FIRST - 3}", "Datenbasis für die Sparklines der KPI-Kacheln",
        "LNP_Abschnitt")
    put(
        ws,
        f"A{KPI_MONAT_FIRST - 2}",
        "Zwölf Monate ab dem ersten Datum der Historie. Die Sparklines legt das "
        "Makro Setup_Arbeitsmappe an.",
        "LNP_Hinweis",
    )
    kopfzeile(
        ws,
        KPI_MONAT_FIRST - 1,
        ["Monat", "bis (exkl.)", "Paletten-Äquivalent", "EUR-Umsatz",
         "Ø Bedarf/Tag (Pal.)", "Bewegungszeilen"],
    )
    datum = rng(HIST, SP_DATUM)
    for i in range(KPI_MONAT_LAST - KPI_MONAT_FIRST + 1):
        r = KPI_MONAT_FIRST + i
        if i == 0:
            put(
                ws,
                f"A{r}",
                f"=IF(COUNT({datum})=0,0,DATE(YEAR(MIN({datum})),MONTH(MIN({datum})),1))",
                "LNP_Datum",
            )
        else:
            put(ws, f"A{r}", f"=DATE(YEAR($A{r - 1}),MONTH($A{r - 1})+1,1)", "LNP_Datum")
        put(ws, f"B{r}", f"=DATE(YEAR($A{r}),MONTH($A{r})+1,1)", "LNP_Datum")
        put(
            ws,
            f"C{r}",
            f"=SUMPRODUCT(({datum}>=$A{r})*({datum}<$B{r})*{rng(HIST, SP_PALAEQ)})",
            "LNP_Zahl2",
        )
        put(
            ws,
            f"D{r}",
            f"=SUMPRODUCT(({datum}>=$A{r})*({datum}<$B{r})*{rng(HIST, SP_UMSATZ)})",
            "LNP_Euro",
        )
        put(ws, f"E{r}", f"=IFERROR($C{r}/($B{r}-$A{r}),0)", "LNP_Zahl2")
        put(ws, f"F{r}", f"=SUMPRODUCT(({datum}>=$A{r})*({datum}<$B{r})*1)", "LNP_Zahl")

    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# 8. Beispieldaten
# ---------------------------------------------------------------------------


def blatt_beispieldaten(wb, hist: list[list], fc: list[list]):
    ws = wb.create_sheet("Beispieldaten")
    titel(
        ws,
        "Beispieldaten (Rohdaten wie aus einem Vorsystem)",
        "Diese Blöcke haben absichtlich andere Spaltennamen als die Import-Blätter, "
        "damit sich die Mapping-Tabelle testen lässt. Über die Schaltfläche "
        '"Beispieldaten als CSV" entsteht daraus eine echte Quelldatei für '
        "Power Query.",
    )
    breiten(ws, {"A": 15, "B": 22, "C": 14, "D": 19, "E": 13, "F": 15, "G": 13,
                 "H": 15, "J": 15, "K": 22, "L": 15, "M": 19, "N": 14, "O": 16,
                 "P": 15, "Q": 14})

    put(ws, "A4", "Historie (Rohformat)", "LNP_Abschnitt")
    kopfzeile(ws, 5, bd.ROH_SPALTEN_HISTORIE)
    for i, zeile in enumerate(hist):
        r = 6 + i
        put(ws, f"A{r}", zeile[0], "LNP_Datum")
        for j, wert in enumerate(zeile[1:], start=1):
            stil = "LNP_Text" if j <= 3 else "LNP_Zahl2"
            put(ws, f"{get_column_letter(1 + j)}{r}", wert, stil)

    put(ws, "J4", "Forecast (Rohformat)", "LNP_Abschnitt")
    kopfzeile(ws, 5, bd.ROH_SPALTEN_FORECAST, first_col=10)
    for i, zeile in enumerate(fc):
        r = 6 + i
        put(ws, f"J{r}", zeile[0], "LNP_Datum")
        for j, wert in enumerate(zeile[1:], start=1):
            stil = "LNP_Text" if j <= 3 else "LNP_Zahl2"
            put(ws, f"{get_column_letter(10 + j)}{r}", wert, stil)
    return ws


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------


def baue(ziel: Path) -> Path:
    hist = bd.historie()
    fc = bd.forecast()

    wb = Workbook()
    wb.remove(wb.active)
    dsg.register(wb)

    blatt_dc(wb)
    blatt_import(
        wb, HIST, "Import – historische Bewegungsdaten",
        "Zielstruktur des Power-Query-Imports. Vorbelegt mit Beispieldaten, damit "
        "die Mappe sofort rechnet; der Import überschreibt diese Tabelle.",
        hist, "tblHistorie",
    )
    blatt_import(
        wb, FC, "Import – Forecast-Daten (zukünftige Perioden)",
        "Gleicher Aufbau wie die Historie, jedoch für extern erstellte "
        "Forecast-Daten mit eigener Quelldatei und eigenem Import-Button.",
        fc, "tblForecast",
    )
    blatt_mapping(wb)
    blatt_zielreichweite(wb)
    blatt_simulation(wb)
    blatt_szenarien(wb)
    blatt_karte(wb)
    blatt_beispieldaten(wb, hist, fc)
    blatt_dashboard(wb)

    # Namen fuer Power Query und VBA
    from openpyxl.workbook.defined_name import DefinedName

    for name, ref in [
        ("Pfad_Historie", "Import_Mapping!$B$4"),
        ("Pfad_Forecast", "Import_Mapping!$B$5"),
        ("Param_Trennzeichen", "Import_Mapping!$B$6"),
        ("Param_M3_pro_Palette", "Import_Mapping!$B$7"),
        ("Param_Quelltabelle", "Import_Mapping!$B$8"),
    ]:
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    wb.properties.title = "Logistik-Netzwerkplanung"
    wb.properties.creator = "Logistik-Netzwerkplanung (Build-Skript)"

    tmp = ziel.with_suffix(".tmp.xlsx")
    wb.save(tmp)

    module = xlsm.load_vba_modules(Path(__file__).parent / "vba")
    projekt = ovba.build_vba_project(module)
    xlsm.convert(tmp, ziel, projekt)
    tmp.unlink()

    print(f"{ziel} geschrieben ({ziel.stat().st_size / 1024:.0f} KB)")
    print(f"  Blätter        : {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
    print(f"  Historie/Forecast: {len(hist)} / {len(fc)} Zeilen")
    print(f"  VBA-Module     : {', '.join(n for n, _ in module)}")
    return ziel


if __name__ == "__main__":
    ausgabe = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Logistik_Netzwerkplanung.xlsm")
    baue(ausgabe)
